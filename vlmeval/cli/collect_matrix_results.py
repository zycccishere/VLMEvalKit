#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path
from typing import Any

import pandas as pd
import yaml


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    if not isinstance(data, dict):
        raise ValueError(f"Config must be a mapping: {path}")
    return data


def format_value(value: Any, repo_root: str) -> Any:
    if isinstance(value, str):
        return value.format(repo_root=repo_root)
    if isinstance(value, list):
        return [format_value(v, repo_root) for v in value]
    if isinstance(value, dict):
        return {k: format_value(v, repo_root) for k, v in value.items()}
    return value


def split_names(raw: str) -> list[str]:
    if not raw:
        return []
    return [part for part in raw.replace(",", " ").split() if part]


def load_task_manifest_rows(path: Path) -> list[dict[str, str]]:
    lower = path.name.lower()
    rows: list[dict[str, str]] = []
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            for raw in csv.DictReader(fh, delimiter=delimiter):
                rows.append({str(k): str(v).strip() for k, v in raw.items() if k and str(v).strip() != ""})
        return rows
    if lower.endswith(".jsonl"):
        with path.open("r", encoding="utf-8") as fh:
            for line in fh:
                text = line.strip()
                if not text:
                    continue
                payload = json.loads(text)
                if not isinstance(payload, dict):
                    raise ValueError(f"Invalid manifest row in {path}: {text[:80]}")
                rows.append({str(k): str(v).strip() for k, v in payload.items() if str(v).strip() != ""})
        return rows
    if lower.endswith(".json"):
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(f"Task manifest JSON must be a list: {path}")
        for item in payload:
            if not isinstance(item, dict):
                raise ValueError(f"Invalid manifest row in {path}: {item!r}")
            rows.append({str(k): str(v).strip() for k, v in item.items() if str(v).strip() != ""})
        return rows
    raise ValueError(f"Unsupported task manifest format: {path}")


def row_matches_manifest(row: dict[str, Any], manifest_rows: list[dict[str, str]]) -> bool:
    aliases = {
        "model": str(row["model_key"]),
        "model_key": str(row["model_key"]),
        "policy": str(row["policy"]),
        "policy_key": str(row["policy"]),
        "mode": str(row["mode"]),
        "transform": str(row["transform"]),
        "dataset": str(row["dataset"]),
    }
    supported = set(aliases.keys())
    for manifest_row in manifest_rows:
        filtered = {key: value for key, value in manifest_row.items() if key in supported}
        if filtered and all(aliases[key] == value for key, value in filtered.items()):
            return True
    return False


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().strip('"')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def maybe_pct(value: float) -> float:
    return value * 100.0 if abs(value) <= 1.0 else value


def choose_overall_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    preferred = {"overall", "average", "none"}
    for row in rows:
        for value in row.values():
            if str(value).strip().lower() in preferred:
                return row
    return rows[0]


def load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    lower = path.name.lower()
    if lower.endswith(".tsv") or lower.endswith(".csv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh, delimiter=delimiter))
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(v) if v is not None else "" for v in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for idx, col in enumerate(header):
                if col:
                    item[col] = row[idx] if idx < len(row) else None
            out.append(item)
        return out
    raise ValueError(f"Unsupported tabular file: {path}")


def parse_csv_metric(path: Path, dataset: str) -> tuple[float, int] | None:
    with path.open("r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        return None
    lower_headers = {key.lower(): key for key in rows[0].keys()}

    if "dataset" in lower_headers and "overall" in lower_headers:
        first_dataset = str(rows[0][lower_headers["dataset"]]).strip()
        value = to_float(rows[0][lower_headers["overall"]])
        if first_dataset == dataset and value is not None:
            return value, 10

    if "setting" in lower_headers and "overall" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["overall"]])
        if value is not None:
            return maybe_pct(value), 30

    if "overall" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["overall"]])
        if value is not None:
            return maybe_pct(value), 20

    if "acc" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["acc"]])
        if value is not None:
            return maybe_pct(value), 20

    return None


def parse_json_metric(path: Path, dataset: str) -> tuple[float, int] | None:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if dataset == "OCRBench":
        value = to_float(data.get("Final Score Norm"))
        if value is not None:
            return value, 30
    overall = to_float(data.get("Overall"))
    if overall is not None:
        return maybe_pct(overall), 30
    return None


def parse_metric_file(path: Path, dataset: str) -> tuple[float, int] | None:
    if path.suffix == ".json":
        return parse_json_metric(path, dataset)
    if path.suffix == ".csv":
        return parse_csv_metric(path, dataset)
    return None


def normalize_index(raw_idx: Any) -> Any:
    if raw_idx is None:
        return None
    try:
        if pd.isna(raw_idx):
            return None
    except Exception:
        pass
    if hasattr(raw_idx, "item"):
        try:
            raw_idx = raw_idx.item()
        except Exception:
            pass
    if isinstance(raw_idx, str):
        stripped = raw_idx.strip()
        if stripped == "":
            return stripped
        try:
            return int(stripped)
        except Exception:
            try:
                maybe_float = float(stripped)
            except Exception:
                return stripped
            if maybe_float.is_integer():
                return int(maybe_float)
            return stripped
    if isinstance(raw_idx, float) and raw_idx.is_integer():
        return int(raw_idx)
    return raw_idx


def load_allowlist(path: str | None) -> set[Any] | None:
    if not path:
        return None
    allowlist_path = Path(path)
    if not allowlist_path.exists():
        raise FileNotFoundError(f"Allowlist file not found: {allowlist_path}")
    if allowlist_path.suffix.lower() == ".json":
        payload = json.loads(allowlist_path.read_text(encoding="utf-8"))
        values = payload.get("indices", []) if isinstance(payload, dict) else payload
    else:
        values = [line.strip() for line in allowlist_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if not isinstance(values, list):
        raise ValueError(f"Invalid allowlist payload in {allowlist_path}")
    out = set()
    for value in values:
        normalized = normalize_index(value)
        if normalized is not None:
            out.add(normalized)
    return out


def filter_data_by_allowlist(data: pd.DataFrame, allowlist: set[Any] | None) -> pd.DataFrame:
    if allowlist is None:
        return data
    normalized = data["index"].map(normalize_index)
    return data[normalized.isin(allowlist)]


def get_expected_count(dataset_name: str, allowlist_path: str | None) -> int:
    from vlmeval.dataset import build_dataset

    dataset = build_dataset(dataset_name)
    if dataset is None:
        return -1
    data = getattr(dataset, "data", None)
    if data is None:
        return int(len(dataset))
    allowlist = load_allowlist(allowlist_path)
    data = filter_data_by_allowlist(data, allowlist)
    return int(len(data))


def infer_complete(pred_file: Path | None, expected: int) -> bool:
    if pred_file is None or expected < 0 or not pred_file.exists():
        return False
    try:
        rows = load_tabular_rows(pred_file)
    except Exception:
        return False
    if len(rows) != expected:
        return False
    candidate_cols = [col for col in ("prediction", "description", "detailed_prediction") if rows and col in rows[0]]
    if not candidate_cols:
        return True
    for row in rows:
        values = [row.get(col) for col in candidate_cols]
        if all(is_blank(value) for value in values):
            return False
        desc = str(row.get("description", "")).strip()
        if desc.startswith("[FAILED_INFER]") or "Failed to obtain answer via API." in desc:
            return False
    return True


def eval_complete(metric_file: Path | None) -> bool:
    if metric_file is None or not metric_file.exists():
        return False
    try:
        return metric_file.stat().st_size > 0
    except OSError:
        return False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect matrix-shaped result tables from a unified-runner result root.")
    parser.add_argument("--matrix-config", type=Path, required=True)
    parser.add_argument("--model-config", type=Path, default=Path("configs/models.yaml"))
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--results-root", type=Path, default=None, help="Optional override for results_root.")
    parser.add_argument("--models", type=str, default="")
    parser.add_argument("--policies", type=str, default="")
    parser.add_argument("--modes", type=str, default="")
    parser.add_argument("--transforms", type=str, default="")
    parser.add_argument("--datasets", type=str, default="")
    parser.add_argument("--task-manifest", type=Path, default=None)
    return parser.parse_args()


def task_root(results_root: Path, explicit_transform_axis: bool, policy: str, mode: str, transform: str, model_key: str) -> Path:
    if explicit_transform_axis:
        return results_root / policy / mode / transform / model_key
    return results_root / policy / mode / model_key


def choose_best_metric_file(model_output_root: Path, dataset: str) -> tuple[Path | None, float | None, str]:
    candidates: list[tuple[int, str, Path, float]] = []
    for path in sorted(model_output_root.glob(f"*{dataset}*_acc.csv")):
        parsed = parse_metric_file(path, dataset)
        if parsed is not None:
            value, priority = parsed
            candidates.append((priority, str(path), path, value))
    for path in sorted(model_output_root.glob(f"*{dataset}*_score.csv")):
        parsed = parse_metric_file(path, dataset)
        if parsed is not None:
            value, priority = parsed
            candidates.append((priority, str(path), path, value))
    for path in sorted(model_output_root.glob(f"*{dataset}*_score.json")):
        parsed = parse_metric_file(path, dataset)
        if parsed is not None:
            value, priority = parsed
            candidates.append((priority, str(path), path, value))
    if not candidates:
        return None, None, ""
    _, _, best_path, best_value = max(candidates, key=lambda item: (item[0], item[1]))
    return best_path, best_value, best_path.suffix.lstrip(".")


def main() -> int:
    args = parse_args()
    matrix = load_yaml(args.matrix_config)
    repo_root = str(matrix["repo_root"])
    matrix = format_value(matrix, repo_root)
    models_cfg = load_yaml(args.model_config)

    results_root = args.results_root or Path(matrix["results_root"])
    explicit_transform_axis = "image_transforms" in matrix
    dataset_allowlists = matrix.get("dataset_index_allowlists", {})

    model_filters = set(split_names(args.models))
    policy_filters = set(split_names(args.policies))
    mode_filters = set(split_names(args.modes))
    transform_filters = set(split_names(args.transforms))
    dataset_filters = set(split_names(args.datasets))

    models = [m for m in matrix["models"] if not model_filters or m in model_filters]
    policies = [p for p in matrix["policies"].keys() if not policy_filters or p in policy_filters]
    modes = [m for m in matrix["replay_modes"] if not mode_filters or m in mode_filters]
    transforms = [t for t in matrix.get("image_transforms", ["baseline"]) if not transform_filters or t in transform_filters]
    datasets = [d for d in matrix["datasets"] if not dataset_filters or d in dataset_filters]
    manifest_rows = load_task_manifest_rows(args.task_manifest) if args.task_manifest is not None else []

    expected_counts = {
        dataset: get_expected_count(dataset, dataset_allowlists.get(dataset))
        for dataset in datasets
    }

    rows: list[dict[str, Any]] = []
    for model_key in models:
        registry_name = str(models_cfg["models"][model_key]["registry_name"])
        for policy in policies:
            for mode in modes:
                for transform in transforms:
                    root = task_root(results_root, explicit_transform_axis, policy, mode, transform, model_key)
                    model_output = root / registry_name
                    for dataset in datasets:
                        xlsx = model_output / f"{registry_name}_{dataset}.xlsx"
                        tsv = model_output / f"{registry_name}_{dataset}.tsv"
                        pred_file = xlsx if xlsx.exists() else tsv if tsv.exists() else None
                        metric_file, metric_value, metric_kind = choose_best_metric_file(model_output, dataset)
                        rows.append(
                            {
                                "model_key": model_key,
                                "registry_name": registry_name,
                                "policy": policy,
                                "mode": mode,
                                "transform": transform,
                                "dataset": dataset,
                                "expected_count": expected_counts[dataset],
                                "infer_complete": infer_complete(pred_file, expected_counts[dataset]),
                                "eval_complete": eval_complete(metric_file),
                                "metric_value": metric_value,
                                "metric_kind": metric_kind,
                                "infer_file": str(pred_file) if pred_file is not None else "",
                                "metric_file": str(metric_file) if metric_file is not None else "",
                            }
                        )

    if manifest_rows:
        rows = [row for row in rows if row_matches_manifest(row, manifest_rows)]

    long_df = pd.DataFrame(rows).sort_values(["model_key", "policy", "mode", "transform", "dataset"]).reset_index(drop=True)
    pivot_df = long_df.pivot_table(
        index=["model_key", "policy", "mode", "transform"],
        columns="dataset",
        values="metric_value",
        aggfunc="first",
    ).reset_index()
    infer_status_df = long_df.pivot_table(
        index=["model_key", "policy", "mode", "transform"],
        columns="dataset",
        values="infer_complete",
        aggfunc="first",
    ).reset_index()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    long_path = args.out_dir / "results_long.csv"
    pivot_path = args.out_dir / "results_pivot.csv"
    infer_status_path = args.out_dir / "infer_status_pivot.csv"
    summary_path = args.out_dir / "summary.json"

    long_df.to_csv(long_path, index=False)
    pivot_df.to_csv(pivot_path, index=False)
    infer_status_df.to_csv(infer_status_path, index=False)

    payload = {
        "matrix_config": str(args.matrix_config),
        "results_root": str(results_root),
        "task_manifest": str(args.task_manifest) if args.task_manifest is not None else "",
        "task_count": len(rows),
        "metric_filled_count": int(long_df["metric_value"].notna().sum()) if not long_df.empty else 0,
        "infer_complete_count": int(long_df["infer_complete"].sum()) if not long_df.empty else 0,
        "datasets": datasets,
        "transforms": transforms,
        "models": models,
        "outputs": {
            "results_long_csv": str(long_path),
            "results_pivot_csv": str(pivot_path),
            "infer_status_pivot_csv": str(infer_status_path),
        },
    }
    summary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
