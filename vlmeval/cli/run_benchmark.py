#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import pickle
import queue
import shutil
import time
import subprocess
import sys
import threading
from collections import Counter, defaultdict
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    import yaml
except ImportError as exc:
    raise SystemExit("PyYAML is required to run vlmeval.cli.run_benchmark") from exc


def truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def contiguous_partition_refines(
    refinement: list[int],
    coarser: list[int],
) -> bool:
    """Whether worker calls split, but never cross, driver call boundaries."""
    if (
        not refinement
        or not coarser
        or any(value <= 0 for value in refinement)
        or any(value <= 0 for value in coarser)
        or sum(refinement) != sum(coarser)
    ):
        return False
    refinement_boundaries = set()
    offset = 0
    for value in refinement:
        offset += value
        refinement_boundaries.add(offset)
    offset = 0
    for value in coarser:
        offset += value
        if offset not in refinement_boundaries:
            return False
    return True


def detect_node_rank() -> int:
    for name in ("NODE_RANK", "SLURM_NODEID", "RANK", "OMPI_COMM_WORLD_RANK", "PMI_RANK"):
        raw = os.environ.get(name)
        if raw not in (None, ""):
            return int(raw)
    return 0


def detect_num_nodes() -> int:
    for name in ("NUM_NODES", "SLURM_NNODES"):
        raw = os.environ.get(name)
        if raw not in (None, ""):
            return int(raw)
    return 1


def split_names(raw: str) -> list[str]:
    if not raw:
        return []
    return [part for part in raw.replace(",", " ").split() if part]


def load_repo_env(repo_root: str) -> None:
    env_path = Path(repo_root) / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise SystemExit(f"{name} is not set. Export {name} or define it in the repo .env before running vlmeval.cli.run_benchmark.")
    return value


DATASET_ALIASES = {
    "MMMU_DEV_VAL_SINGLE": "MMMU_DEV_VAL_SINGLE_IMAGE",
}


def canonical_dataset_name(name: str) -> str:
    return DATASET_ALIASES.get(str(name), str(name))


def canonical_dataset_names(names: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for name in names:
        canonical = canonical_dataset_name(name)
        if canonical in seen:
            continue
        seen.add(canonical)
        out.append(canonical)
    return out


def load_task_manifest_rows(path: Path) -> list[dict[str, str]]:
    lower = path.name.lower()
    rows: list[dict[str, str]] = []
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            for raw in csv.DictReader(fh, delimiter=delimiter):
                rows.append({str(k): str(v).strip() for k, v in raw.items() if k and str(v).strip() != ""})
        return rows
    if lower.endswith(".jsonl"):
        with path.open("r", encoding="utf-8") as fh:
            for line in fh:
                text = line.strip()
                if not text:
                    continue
                payload = json.loads(text)
                if not isinstance(payload, dict):
                    raise ValueError(f"Invalid manifest row in {path}: {text[:80]}")
                rows.append({str(k): str(v).strip() for k, v in payload.items() if str(v).strip() != ""})
        return rows
    if lower.endswith(".json"):
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(f"Task manifest JSON must be a list: {path}")
        for item in payload:
            if not isinstance(item, dict):
                raise ValueError(f"Invalid manifest row in {path}: {item!r}")
            rows.append({str(k): str(v).strip() for k, v in item.items() if str(v).strip() != ""})
        return rows
    raise ValueError(f"Unsupported task manifest format: {path}")


def task_matches_manifest(task: "Task", rows: list[dict[str, str]]) -> bool:
    aliases = {
        "model": task.model_key,
        "model_key": task.model_key,
        "policy": task.policy_key,
        "policy_key": task.policy_key,
        "mode": task.mode,
        "transform": task.transform,
        "visual_token_shift": task.visual_token_shift,
        "dataset": task.dataset,
    }
    supported = set(aliases.keys())
    for row in rows:
        filtered = {key: value for key, value in row.items() if key in supported}
        if filtered and all(
            aliases[key] == (canonical_dataset_name(value) if key == "dataset" else value)
            for key, value in filtered.items()
        ):
            return True
    return False


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    if not isinstance(data, dict):
        raise ValueError(f"Config must be a mapping: {path}")
    return data


def format_value(value: Any, repo_root: str) -> Any:
    if isinstance(value, str):
        token_roll_user_root = str(Path(repo_root).resolve().parent)
        token_roll_runtime_root = str(Path(token_roll_user_root) / ".venvs")
        mapping = {
            "repo_root": repo_root,
            "model_root": os.environ.get("MODEL_ROOT", "/models"),
            "conda_root": os.environ.get("CONDA_ROOT", os.environ.get("HOME", "") + "/miniconda3"),
            "lmu_data": require_env("LMUData"),
            "token_roll_python": os.environ.get(
                "TOKEN_ROLL_PYTHON",
                str(Path(token_roll_runtime_root) / "lmms-engine/bin/python"),
            ),
            "token_roll_pydeps": os.environ.get(
                "TOKEN_ROLL_PYDEPS",
                str(Path(token_roll_runtime_root) / "vlmevalkit-token-roll-pydeps"),
            ),
            "minicpm_token_roll_pydeps": os.environ.get(
                "MINICPM_TOKEN_ROLL_PYDEPS",
                str(Path(token_roll_runtime_root) / "minicpmo-token-roll-pydeps"),
            ),
            "qwen_token_roll_vllm_python": os.environ.get(
                "QWEN_TOKEN_ROLL_VLLM_PYTHON",
                "/user/wanzihao/miniconda3/envs/vlmevalkit/bin/python",
            ),
            "minicpm_token_roll_vllm_python": os.environ.get(
                "MINICPM_TOKEN_ROLL_VLLM_PYTHON",
                "/user/zhangyicheng/miniconda3/envs/vlmeval_qwen35_vllm/bin/python",
            ),
            "vllm_token_roll_plugin_overlay": os.environ.get(
                "VLLM_TOKEN_ROLL_PLUGIN_OVERLAY",
                str(Path(repo_root) / ".vllm_plugin_overlay"),
            ),
        }
        try:
            return value.format(**mapping)
        except (KeyError, ValueError):
            # Historical matrix files may contain literal braces. Keep those
            # strings unchanged rather than failing before task construction.
            return value
    if isinstance(value, list):
        return [format_value(v, repo_root) for v in value]
    if isinstance(value, dict):
        return {k: format_value(v, repo_root) for k, v in value.items()}
    return value


def prepend_pythonpath(env: dict[str, str], entries: list[str]) -> None:
    cleaned = [entry for entry in entries if entry]
    if not cleaned:
        return
    existing = env.get("PYTHONPATH", "")
    parts = cleaned[:]
    if existing:
        parts.append(existing)
    env["PYTHONPATH"] = ":".join(parts)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    lower = path.name.lower()
    if lower.endswith(".tsv") or lower.endswith(".csv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh, delimiter=delimiter))
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(v) if v is not None else "" for v in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for idx, col in enumerate(header):
                if col:
                    item[col] = row[idx] if idx < len(row) else None
            out.append(item)
        return out
    raise ValueError(f"Unsupported tabular file: {path}")


@dataclass(frozen=True)
class EnvProfile:
    key: str
    python: str
    pythonpath: tuple[str, ...]
    base_env: dict[str, str]
    ensure_modules: tuple[str, ...]


@dataclass(frozen=True)
class ModelSpec:
    key: str
    display_name: str
    registry_name: str
    env_profile: str
    model_path: str
    gpus_per_job: int
    infer_batch_size: int
    max_num_seqs: int
    tp_size: int
    max_model_len: int | None
    estimated_dataset_cost: float
    task_env: dict[str, str]

    @property
    def workers_per_node(self) -> int:
        return 8 // self.gpus_per_job

    @property
    def node_time_weight(self) -> float:
        return self.estimated_dataset_cost / self.workers_per_node


@dataclass(frozen=True)
class PolicySpec:
    key: str
    replay_prompt_template_name: str


@dataclass(frozen=True)
class Task:
    index: int
    model_key: str
    policy_key: str
    mode: str
    transform: str
    visual_token_shift: str
    dataset: str
    weight: float

    @property
    def tag(self) -> str:
        parts = [self.model_key, self.policy_key, self.mode, self.transform]
        if self.visual_token_shift != "none":
            parts.append(self.visual_token_shift)
        parts.append(self.dataset)
        return "__".join(parts)


class BenchmarkRunner:
    def __init__(self, repo_root: Path, args: argparse.Namespace) -> None:
        self.repo_root_path = repo_root
        self.args = args
        self.matrix_config_path = Path(args.matrix_config).resolve()
        self.model_config_path = Path(args.model_config).resolve()
        matrix = load_yaml(args.matrix_config)
        models_cfg = load_yaml(args.model_config)
        default_repo_root = str(repo_root)
        load_repo_env(default_repo_root)
        self.repo_root = str(format_value(matrix.get("repo_root", default_repo_root), default_repo_root))
        self.results_root = self._resolve_results_root(matrix["results_root"])
        self.matrix_name = matrix["name"]
        self.node_gpu_ids = split_names(args.gpu_ids or str(matrix.get("node_gpu_ids", "0,1,2,3,4,5,6,7")))
        self.resume_infer = args.resume_infer if args.resume_infer is not None else truthy(matrix.get("resume_infer_default", False))
        self.inference_cfg = matrix.get("inference", {})
        self.inference_launch_mode = str(
            self.inference_cfg.get("launch_mode", "run")
        ).strip().lower()
        if self.inference_launch_mode not in {"run", "existing"}:
            raise ValueError(
                f"Unsupported inference launch mode: {self.inference_launch_mode}"
            )
        self.inference_producer_matrix = str(
            self.inference_cfg.get("producer_matrix", "")
        ).strip()
        if self.inference_launch_mode == "existing" and not self.inference_producer_matrix:
            raise ValueError(
                "inference.launch_mode=existing requires inference.producer_matrix"
            )
        self.evaluation_cfg = matrix.get("evaluation", {})
        self.replay_cfg = matrix.get("replay", {})
        self.answer_format_cfg = matrix.get("answer_format", {})
        self.trace_cfg = matrix.get("trace", {})
        self.infer_batch_size_override = (
            int(matrix["infer_batch_size_override"])
            if matrix.get("infer_batch_size_override") is not None
            else None
        )
        self.max_num_seqs_override = (
            int(matrix["max_num_seqs_override"])
            if matrix.get("max_num_seqs_override") is not None
            else None
        )
        self.scheduler = str(args.scheduler or matrix.get("scheduler", "model_sequential")).strip()
        if self.scheduler not in {"model_sequential", "gpu_pool"}:
            raise ValueError(f"Unsupported scheduler: {self.scheduler}")
        raw_allowlists = format_value(matrix.get("dataset_index_allowlists", {}), self.repo_root)
        self.dataset_index_allowlists = {
            canonical_dataset_name(str(dataset)): path for dataset, path in raw_allowlists.items()
        }
        self.explicit_transform_axis = "image_transforms" in matrix
        self.explicit_visual_token_shift_axis = "visual_token_shifts" in matrix
        self.worker_monitor_cfg = matrix.get("worker_monitor", {})
        self.worker_monitor_enabled = truthy(self.worker_monitor_cfg.get("enable", True))
        interval_raw = os.environ.get("WORKER_MONITOR_INTERVAL_SECONDS", self.worker_monitor_cfg.get("interval_seconds", 600))
        self.worker_monitor_interval = max(1, int(interval_raw))
        resume_delay_raw = os.environ.get(
            "RESUME_INFER_START_DELAY_SECONDS",
            matrix.get("resume_infer_start_delay_seconds", 10),
        )
        self.resume_infer_start_delay_seconds = max(0, int(resume_delay_raw))
        self.expected_count_cache: dict[tuple[str, str], int] = {}
        self.content_identity_cache: dict[str, tuple[tuple[Any, ...], dict[str, Any]]] = {}
        self.content_identity_lock = threading.Lock()
        self.env_ready: set[str] = set()
        self.env_ready_lock = threading.Lock()
        self.failures: list[str] = []
        self.failure_lock = threading.Lock()
        self.task_manifest_path = args.task_manifest
        self.task_manifest_rows: list[dict[str, str]] = []
        self._load_profiles(models_cfg)
        self._load_matrix(matrix, models_cfg)

    def _resolve_results_root(self, raw: str) -> Path:
        path = Path(raw)
        if path.is_absolute():
            return path
        return Path(self.repo_root) / raw

    def _load_profiles(self, models_cfg: dict[str, Any]) -> None:
        self.env_profiles: dict[str, EnvProfile] = {}
        for key, raw in models_cfg["env_profiles"].items():
            resolved = format_value(raw, self.repo_root)
            self.env_profiles[key] = EnvProfile(
                key=key,
                python=str(resolved["python"]),
                pythonpath=tuple(str(v) for v in resolved.get("pythonpath", [])),
                base_env={k: str(v) for k, v in resolved.get("base_env", {}).items()},
                ensure_modules=tuple(str(v) for v in resolved.get("ensure_modules", [])),
            )

    def _load_matrix(self, matrix: dict[str, Any], models_cfg: dict[str, Any]) -> None:
        policy_filters = split_names(self.args.policies)
        mode_filters = split_names(self.args.modes)
        transform_filters = split_names(self.args.transforms)
        visual_token_shift_filters = split_names(self.args.visual_token_shifts)
        dataset_filters = canonical_dataset_names(split_names(self.args.datasets))
        model_filters = split_names(self.args.models)

        policy_order = list(matrix["policies"].keys())
        mode_order = list(matrix["replay_modes"])
        transform_order = list(matrix.get("image_transforms", ["baseline"]))
        visual_token_shift_order = list(matrix.get("visual_token_shifts", ["none"]))
        dataset_order = canonical_dataset_names([str(name) for name in matrix["datasets"]])
        model_order = list(matrix["models"])

        self.policy_order = [name for name in policy_order if not policy_filters or name in policy_filters]
        self.mode_order = [name for name in mode_order if not mode_filters or name in mode_filters]
        self.transform_order = [name for name in transform_order if not transform_filters or name in transform_filters]
        self.visual_token_shift_order = [
            name
            for name in visual_token_shift_order
            if not visual_token_shift_filters or name in visual_token_shift_filters
        ]
        self.dataset_order = [name for name in dataset_order if not dataset_filters or name in dataset_filters]
        self.model_order = [name for name in model_order if not model_filters or name in model_filters]

        self.policies = {
            name: PolicySpec(key=name, replay_prompt_template_name=str(matrix["policies"][name]["replay_prompt_template_name"]))
            for name in self.policy_order
        }
        self.models: dict[str, ModelSpec] = {}
        for key in self.model_order:
            raw = format_value(models_cfg["models"][key], self.repo_root)
            runtime = raw["runtime"]
            self.models[key] = ModelSpec(
                key=key,
                display_name=str(raw["display_name"]),
                registry_name=str(raw["registry_name"]),
                env_profile=str(raw["env_profile"]),
                model_path=str(raw["model_path"]),
                gpus_per_job=int(runtime["gpus_per_job"]),
                infer_batch_size=int(runtime["infer_batch_size"]),
                max_num_seqs=int(runtime["max_num_seqs"]),
                tp_size=int(runtime["tp_size"]),
                max_model_len=int(runtime["max_model_len"]) if runtime.get("max_model_len") is not None else None,
                estimated_dataset_cost=float(runtime["estimated_dataset_cost"]),
                task_env={k: str(v) for k, v in raw.get("task_env", {}).items()},
            )

        self.tasks: list[Task] = []
        index = 0
        for model_key in self.model_order:
            model = self.models[model_key]
            for policy_key in self.policy_order:
                for mode in self.mode_order:
                    for transform in self.transform_order:
                        for visual_token_shift in self.visual_token_shift_order:
                            for dataset in self.dataset_order:
                                self.tasks.append(
                                    Task(
                                        index=index,
                                        model_key=model_key,
                                        policy_key=policy_key,
                                        mode=mode,
                                        transform=transform,
                                        visual_token_shift=visual_token_shift,
                                        dataset=dataset,
                                        weight=model.node_time_weight,
                                    )
                                )
                                index += 1
        if self.task_manifest_path is not None:
            self.task_manifest_rows = load_task_manifest_rows(self.task_manifest_path)
            self.tasks = [task for task in self.tasks if task_matches_manifest(task, self.task_manifest_rows)]

    def plan(self) -> list[list[Task]]:
        if self.args.manifest_is_node_shard:
            if self.task_manifest_path is None:
                raise SystemExit("--manifest-is-node-shard requires --task-manifest")
            if self.args.node_rank < 0 or self.args.node_rank >= self.args.nodes:
                raise SystemExit(f"node-rank out of range: {self.args.node_rank} / {self.args.nodes}")
            buckets: list[list[Task]] = [[] for _ in range(self.args.nodes)]
            buckets[self.args.node_rank] = sorted(self.tasks, key=lambda task: task.index)
            self.planned_loads = [0.0 for _ in range(self.args.nodes)]
            self.planned_loads[self.args.node_rank] = sum(task.weight for task in self.tasks)
            return buckets

        groups: list[dict[str, Any]] = []
        for model_key in self.model_order:
            model_tasks = [task for task in self.tasks if task.model_key == model_key]
            if not model_tasks:
                continue
            groups.append(
                {
                    "model_key": model_key,
                    "tasks": model_tasks,
                    "load": len(model_tasks) * self.models[model_key].node_time_weight,
                }
            )

        if not groups:
            self.planned_loads = [0.0 for _ in range(self.args.nodes)]
            return [[] for _ in range(self.args.nodes)]

        while len(groups) < self.args.nodes:
            split_idx = max(
                range(len(groups)),
                key=lambda idx: (groups[idx]["load"], len(groups[idx]["tasks"])),
            )
            group = groups[split_idx]
            if len(group["tasks"]) <= 1:
                break
            midpoint = len(group["tasks"]) // 2
            left_tasks = group["tasks"][:midpoint]
            right_tasks = group["tasks"][midpoint:]
            model = self.models[group["model_key"]]
            left = {
                "model_key": group["model_key"],
                "tasks": left_tasks,
                "load": len(left_tasks) * model.node_time_weight,
            }
            right = {
                "model_key": group["model_key"],
                "tasks": right_tasks,
                "load": len(right_tasks) * model.node_time_weight,
            }
            groups[split_idx : split_idx + 1] = [left, right]

        buckets: list[list[Task]] = [[] for _ in range(self.args.nodes)]
        loads = [0.0 for _ in range(self.args.nodes)]
        ordered_groups = sorted(
            groups,
            key=lambda group: (-group["load"], self.model_order.index(group["model_key"])),
        )
        for group in ordered_groups:
            node_index = min(range(self.args.nodes), key=lambda idx: (loads[idx], idx))
            buckets[node_index].extend(group["tasks"])
            loads[node_index] += group["load"]
        for bucket in buckets:
            bucket.sort(key=lambda task: task.index)
        self.planned_loads = loads
        return buckets

    def print_plan(self, buckets: list[list[Task]]) -> None:
        print(
            json.dumps(
                {
                    "matrix": self.matrix_name,
                    "nodes": self.args.nodes,
                    "node_rank": self.args.node_rank,
                    "gpu_ids": self.node_gpu_ids,
                    "resume_infer": self.resume_infer,
                    "inference_launch_mode": self.inference_launch_mode,
                    "inference_producer_matrix": self.inference_producer_matrix,
                    "scheduler": self.scheduler,
                    "results_root": str(self.results_root),
                    "task_manifest": str(self.task_manifest_path) if self.task_manifest_path is not None else "",
                    "task_count": len(self.tasks),
                    "model_order": self.model_order,
                    "model_runtime": {
                        model_key: {
                            "env_profile": self.models[model_key].env_profile,
                            "gpus_per_job": self.models[model_key].gpus_per_job,
                            "infer_batch_size": self.infer_batch_size_override
                            or self.models[model_key].infer_batch_size,
                            "max_num_seqs": self.max_num_seqs_override
                            or self.models[model_key].max_num_seqs,
                            "tp_size": self.models[model_key].tp_size,
                        }
                        for model_key in self.model_order
                    },
                    "policy_order": self.policy_order,
                    "mode_order": self.mode_order,
                    "transform_order": self.transform_order,
                    "visual_token_shift_order": self.visual_token_shift_order,
                    "dataset_order": self.dataset_order,
                },
                ensure_ascii=False,
                indent=2,
            ),
            flush=True,
        )
        for node_idx, bucket in enumerate(buckets):
            counts = Counter(task.model_key for task in bucket)
            summary = {
                "node": node_idx,
                "load": round(self.planned_loads[node_idx], 4),
                "tasks": len(bucket),
                "models": dict(sorted(counts.items())),
            }
            print(f"[PLAN] {json.dumps(summary, ensure_ascii=False)}", flush=True)

    def run(self) -> int:
        buckets = self.plan()
        self.print_plan(buckets)
        if self.args.plan_only:
            return 0
        if self.args.node_rank < 0 or self.args.node_rank >= self.args.nodes:
            raise SystemExit(f"node-rank out of range: {self.args.node_rank} / {self.args.nodes}")
        assigned = buckets[self.args.node_rank]
        if not assigned:
            print(f"[NODE][IDLE] node_rank={self.args.node_rank} has no assigned tasks.", flush=True)
            return 0
        if self.scheduler == "gpu_pool":
            self.run_gpu_pool(assigned)
            if self.failures:
                print(f"[NODE][FAIL] node_rank={self.args.node_rank} failures={len(self.failures)}", flush=True)
                for failure in self.failures:
                    print(f"[NODE][FAILURE] {failure}", flush=True)
                return 1
            print(f"[NODE][DONE] node_rank={self.args.node_rank}", flush=True)
            return 0

        grouped: dict[str, list[Task]] = defaultdict(list)
        for task in assigned:
            grouped[task.model_key].append(task)

        for model_key in self.model_order:
            if model_key not in grouped:
                continue
            model = self.models[model_key]
            self.ensure_profile_ready(model)
            tasks = sorted(grouped[model_key], key=lambda task: task.index)
            gpu_chunks = self._gpu_chunks(model.gpus_per_job)
            worker_count = min(len(gpu_chunks), len(tasks))
            task_queue: queue.Queue[Task] = queue.Queue()
            for task in tasks:
                task_queue.put(task)
            print(
                f"[NODE][MODEL] node_rank={self.args.node_rank} model={model_key} "
                f"workers={worker_count} gpus_per_job={model.gpus_per_job} tasks={len(tasks)} scheduling=dynamic",
                flush=True,
            )
            with ThreadPoolExecutor(max_workers=worker_count) as pool:
                futures = []
                for worker_idx in range(worker_count):
                    futures.append(
                        pool.submit(
                            self.run_worker,
                            model,
                            worker_idx,
                            gpu_chunks[worker_idx],
                            task_queue,
                        )
                    )
                for future in futures:
                    future.result()
        if self.failures:
            print(f"[NODE][FAIL] node_rank={self.args.node_rank} failures={len(self.failures)}", flush=True)
            for failure in self.failures:
                print(f"[NODE][FAILURE] {failure}", flush=True)
            return 1
        print(f"[NODE][DONE] node_rank={self.args.node_rank}", flush=True)
        return 0

    def run_gpu_pool(self, assigned: list[Task]) -> None:
        for task in assigned:
            model = self.models[task.model_key]
            if model.gpus_per_job > len(self.node_gpu_ids):
                raise SystemExit(
                    f"Task {task.tag} requires {model.gpus_per_job} GPUs, "
                    f"but node has only {len(self.node_gpu_ids)} GPU ids: {self.node_gpu_ids}"
                )

        for model_key in self.model_order:
            if any(task.model_key == model_key for task in assigned):
                self.ensure_profile_ready(self.models[model_key])

        gpu_order = {gpu: idx for idx, gpu in enumerate(self.node_gpu_ids)}
        available_gpus = list(self.node_gpu_ids)
        pending = sorted(
            assigned,
            key=lambda task: (
                -self.models[task.model_key].gpus_per_job,
                -task.weight,
                self.model_order.index(task.model_key),
                task.index,
            ),
        )
        running_model_counts: Counter[str] = Counter()
        running: dict[Any, tuple[Task, ModelSpec, list[str]]] = {}
        launch_count = 0
        completed_count = 0
        print(
            f"[NODE][SCHEDULER] node_rank={self.args.node_rank} scheduler=gpu_pool "
            f"gpus={','.join(self.node_gpu_ids)} tasks={len(assigned)}",
            flush=True,
        )
        with ThreadPoolExecutor(max_workers=max(1, len(self.node_gpu_ids))) as pool:
            while pending or running:
                while pending and available_gpus:
                    task_idx = self._select_gpu_pool_task(pending, len(available_gpus), running_model_counts)
                    if task_idx is None:
                        break
                    task = pending.pop(task_idx)
                    model = self.models[task.model_key]
                    gpu_ids = available_gpus[: model.gpus_per_job]
                    del available_gpus[: model.gpus_per_job]
                    worker_idx = launch_count
                    launch_count += 1
                    running_model_counts[model.key] += 1
                    print(
                        f"[GPU_POOL][LAUNCH] node_rank={self.args.node_rank} worker={worker_idx} "
                        f"task={task.tag} gpus={','.join(gpu_ids)} free={','.join(available_gpus) or '-'}",
                        flush=True,
                    )
                    future = pool.submit(self.run_one_task_worker, model, worker_idx, gpu_ids, task)
                    running[future] = (task, model, gpu_ids)

                if not running:
                    if pending:
                        next_task = pending[0]
                        next_model = self.models[next_task.model_key]
                        raise RuntimeError(
                            f"No schedulable task with free GPUs={available_gpus}; "
                            f"next={next_task.tag} requires {next_model.gpus_per_job}"
                        )
                    break

                done, _ = wait(list(running.keys()), return_when=FIRST_COMPLETED)
                for future in done:
                    task, model, gpu_ids = running.pop(future)
                    try:
                        future.result()
                    except Exception as exc:
                        self.record_failure(task, f"{type(exc).__name__}: {exc}")
                        print(f"[GPU_POOL][FAIL] {task.tag}: {type(exc).__name__}: {exc}", flush=True)
                    available_gpus.extend(gpu_ids)
                    available_gpus.sort(key=lambda gpu: gpu_order[gpu])
                    running_model_counts[model.key] -= 1
                    if running_model_counts[model.key] <= 0:
                        del running_model_counts[model.key]
                    completed_count += 1
                    print(
                        f"[GPU_POOL][RELEASE] node_rank={self.args.node_rank} task={task.tag} "
                        f"gpus={','.join(gpu_ids)} completed={completed_count}/{len(assigned)} "
                        f"free={','.join(available_gpus)}",
                        flush=True,
                    )

    def _select_gpu_pool_task(
        self,
        pending: list[Task],
        free_gpu_count: int,
        running_model_counts: Counter[str],
    ) -> int | None:
        candidates = [
            (idx, task)
            for idx, task in enumerate(pending)
            if self.models[task.model_key].gpus_per_job <= free_gpu_count
        ]
        if not candidates:
            return None

        # Prefer the packing requested for mixed large-model nodes:
        # when a 4-GPU task is already running and four GPUs remain, choose a
        # 2-GPU task first so the pool can reach 1xTP4 + 2xTP2 instead of
        # immediately launching another TP4 job.
        running_large = any(self.models[model_key].gpus_per_job >= 4 for model_key in running_model_counts)
        if running_large and free_gpu_count == 4:
            two_gpu = [
                (idx, task)
                for idx, task in candidates
                if self.models[task.model_key].gpus_per_job == 2
            ]
            if two_gpu:
                return min(
                    two_gpu,
                    key=lambda item: (
                        -item[1].weight,
                        self.model_order.index(item[1].model_key),
                        item[1].index,
                    ),
                )[0]

        return min(
            candidates,
            key=lambda item: (
                -self.models[item[1].model_key].gpus_per_job,
                -item[1].weight,
                self.model_order.index(item[1].model_key),
                item[1].index,
            ),
        )[0]

    def run_one_task_worker(self, model: ModelSpec, worker_idx: int, gpu_ids: list[str], task: Task) -> None:
        one_task_queue: queue.Queue[Task] = queue.Queue()
        one_task_queue.put(task)
        self.run_worker(model, worker_idx, gpu_ids, one_task_queue)

    def _gpu_chunks(self, gpus_per_job: int) -> list[list[str]]:
        if len(self.node_gpu_ids) % gpus_per_job != 0:
            raise SystemExit(f"GPU list {self.node_gpu_ids} is not divisible by gpus_per_job={gpus_per_job}")
        return [
            self.node_gpu_ids[idx : idx + gpus_per_job]
            for idx in range(0, len(self.node_gpu_ids), gpus_per_job)
        ]

    def ensure_profile_ready(self, model: ModelSpec) -> None:
        profile = self.env_profiles[model.env_profile]
        with self.env_ready_lock:
            if profile.key in self.env_ready:
                return
        if not Path(profile.python).exists():
            raise SystemExit(f"python not found for env profile {profile.key}: {profile.python}")
        for module_name in profile.ensure_modules:
            check = subprocess.run(
                [profile.python, "-c", f"import {module_name}"],
                cwd=self.repo_root,
                env=self._base_env_for_profile(profile),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            if check.returncode == 0:
                continue
            print(f"[SETUP][INSTALL] env={profile.key} missing module={module_name}, installing from Tsinghua mirror", flush=True)
            install = subprocess.run(
                [
                    profile.python,
                    "-m",
                    "pip",
                    "install",
                    "-i",
                    "https://pypi.tuna.tsinghua.edu.cn/simple",
                    "--no-user",
                    module_name,
                ],
                cwd=self.repo_root,
                env=self._base_env_for_profile(profile),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
            )
            if install.returncode != 0:
                raise RuntimeError(f"failed to install {module_name} for {profile.key}:\n{install.stdout}")
        with self.env_ready_lock:
            self.env_ready.add(profile.key)

    def _base_env_for_profile(self, profile: EnvProfile) -> dict[str, str]:
        env = dict(os.environ)
        env.update(profile.base_env)
        prepend_pythonpath(env, list(profile.pythonpath))
        return env

    def _is_qwen25vl_model(self, model: ModelSpec) -> bool:
        text = " ".join([model.key, model.display_name, model.registry_name, model.model_path]).lower()
        return "qwen25vl" in text or "qwen2.5-vl" in text

    def _logicvista_qwen25vl_v0_enabled(self, model: ModelSpec, task: Task) -> bool:
        value = str(os.environ.get("LOGICVISTA_QWEN25VL_FORCE_V0", "1")).strip().lower()
        return task.dataset == "LogicVista" and self._is_qwen25vl_model(model) and value not in {"0", "false", "no", "off"}

    def _apply_qwen25vl_sampling_defaults(self, env: dict[str, str], model: ModelSpec, task: Task) -> None:
        if not self._is_qwen25vl_model(model):
            return
        env.setdefault("VLLM_USE_V1", "1")
        sampling_keys = [
            "QWEN2VL_VLLM_REPETITION_PENALTY",
            "QWEN2VL_VLLM_TEMPERATURE",
            "QWEN2VL_VLLM_TOP_P",
            "QWEN2VL_VLLM_TOP_K",
            "QWEN2VL_VLLM_MAX_TOKENS",
            "QWEN2VL_VLLM_STOP_TOKEN_IDS",
            "QWEN2VL_VLLM_SEED",
        ]
        if not self._logicvista_qwen25vl_v0_enabled(model, task):
            for key in sampling_keys:
                env.pop(key, None)
            return
        env["LOGICVISTA_QWEN25VL_FORCE_V0"] = "1"
        env["LOGICVISTA_QWEN25VL_LEGACY_SAMPLING"] = "1"
        env["LOGICVISTA_QWEN25VL_BATCH_SIZE"] = str(
            self.infer_batch_size_for_task(model, task)
        )
        env["LOGICVISTA_QWEN25VL_MAX_NUM_SEQS"] = str(
            self.max_num_seqs_for_task(model, task)
        )
        env["VLLM_USE_V1"] = "0"
        env.setdefault("QWEN2VL_VLLM_REPETITION_PENALTY", "1.05")
        env.setdefault("QWEN2VL_VLLM_TEMPERATURE", "0.01")
        env.setdefault("QWEN2VL_VLLM_TOP_P", "1.0")
        env.setdefault("QWEN2VL_VLLM_TOP_K", "0")
        env.setdefault("QWEN2VL_VLLM_MAX_TOKENS", "2048")
        env.setdefault("QWEN2VL_VLLM_STOP_TOKEN_IDS", "151645,151643")
        env.setdefault("QWEN2VL_VLLM_SEED", "1234")

    def infer_batch_size_for_task(self, model: ModelSpec, task: Task) -> int:
        return self.infer_batch_size_override or model.infer_batch_size

    def max_num_seqs_for_task(self, model: ModelSpec, task: Task) -> int:
        return self.max_num_seqs_override or model.max_num_seqs

    def build_env(self, model: ModelSpec, task: Task, gpu_ids: list[str]) -> dict[str, str]:
        profile = self.env_profiles[model.env_profile]
        policy = self.policies[task.policy_key]
        env = self._base_env_for_profile(profile)
        env.update(model.task_env)
        self._apply_qwen25vl_sampling_defaults(env, model, task)
        judge_api_key = str(
            os.environ.get("OPENAI_API_KEY_JUDGE")
            or os.environ.get("OPENAI_API_KEY")
            or self.evaluation_cfg.get("openai_api_key", "")
        ).strip()
        judge_api_base = str(
            os.environ.get("OPENAI_API_BASE_JUDGE")
            or os.environ.get("OPENAI_API_BASE")
            or self.evaluation_cfg.get("openai_api_base", "")
        ).strip()
        if judge_api_key:
            env["OPENAI_API_KEY"] = judge_api_key
            env["OPENAI_API_KEY_JUDGE"] = judge_api_key
        if judge_api_base:
            env["OPENAI_API_BASE"] = judge_api_base
            env["OPENAI_API_BASE_JUDGE"] = judge_api_base
        if (
            not env.get("VLMEVAL_API_USAGE_LOG_FILE")
            and not env.get("TOKEN_USAGE_LOG_FILE")
            and truthy(env.get("VLMEVAL_API_USAGE_LOG_DEFAULT", "1"))
        ):
            env["VLMEVAL_API_USAGE_LOG_FILE"] = str(self.usage_log_path(task, model))
        env["CUDA_VISIBLE_DEVICES"] = ",".join(gpu_ids)
        env["MODEL_PATH"] = model.model_path
        env["REPLAY_MODE"] = task.mode
        env["REPLAY_IMAGE_TRANSFORM"] = task.transform
        env["REPLAY_IMAGE_TRANSFORM_CACHE_DIR"] = str(self.task_root(task) / "_transform_cache" / task.dataset)
        env["REPLAY_IMAGE_TRANSFORM_TARGET_POSITION"] = "2"
        env["REPLAY_VISUAL_TOKEN_SHIFT"] = task.visual_token_shift
        env["REPLAY_VISUAL_TOKEN_SHIFT_TARGET_POSITION"] = "2"
        env["REPLAY_VISUAL_TOKEN_SHIFT_STRICT"] = "1"
        env["REPLAY_VISUAL_TOKEN_SHIFT_DUMP_SAMPLES"] = str(self.trace_cfg.get("samples", 1))
        env["REPLAY_VISUAL_TOKEN_SHIFT_RAW_DUMP"] = str(
            int(truthy(self.trace_cfg.get("visual_token_raw_dump", False)))
        )
        env["REPLAY_VISUAL_TOKEN_SHIFT_DUMP_DIR"] = str(self.task_root(task) / "_visual_token_shift")
        matched_token_roll_runtime = truthy(
            env.get("REPLAY_VLLM_MATCHED_TOKEN_ROLL_RUNTIME", "0")
        )
        if task.visual_token_shift != "none" or matched_token_roll_runtime:
            if self._is_qwen25vl_model(model):
                env["REPLAY_VLLM_TARGET_FAMILY"] = "qwen2_5_vl"
            elif model.key.startswith("minicpm_o_45"):
                env["REPLAY_VLLM_TARGET_FAMILY"] = "minicpm_o_4_5"
                env["REPLAY_MINICPM_EXPECTED_QUERY_NUM"] = "64"
            else:
                raise ValueError(
                    f"visual-token shift has no vLLM target family for model {model.key}"
                )
        env["REPLAY_PROMPT_TEMPLATE_NAME"] = policy.replay_prompt_template_name
        env["REPLAY_TIMES"] = str(self.replay_cfg.get("replay_times", 1))
        env["REPLAY_DEBUG"] = "0"
        env["REPLAY_IMAGE_COPY_MODE"] = str(self.replay_cfg.get("image_copy_mode", "reuse_path"))
        env["REPLAY_TEMPLATE_ON_LAST_REPLAY_TEXT"] = str(self.replay_cfg.get("template_on_last_replay_text", 1))
        env["REPLAY_LIMIT_MM_PER_PROMPT"] = str(self.replay_cfg.get("limit_mm_per_prompt", 2))
        env["REPLAY_SAFE_FALLBACK"] = str(self.replay_cfg.get("safe_fallback", 0))
        env["REPLAY_SAFE_TRUNCATE_CHARS"] = str(self.replay_cfg.get("safe_truncate_chars", 6000))
        env["REPLAY_STAGE_DEBUG"] = str(self.replay_cfg.get("stage_debug", 1))
        env["REPLAY_STAGE_DEBUG_SAMPLES"] = str(self.replay_cfg.get("stage_debug_samples", 8))
        env["REPLAY_PROMPT_AUDIT"] = str(self.replay_cfg.get("prompt_audit", 1))
        env["REPLAY_PROMPT_AUDIT_PRINT"] = str(self.replay_cfg.get("prompt_audit_print", 1))
        env["VLMEVAL_STRICT_BATCH"] = str(self.replay_cfg.get("strict_batch", 1))
        if "force_common_prompt" in self.replay_cfg:
            env["REPLAY_FORCE_COMMON_PROMPT"] = str(self.replay_cfg.get("force_common_prompt", 0))
        trace_level = str(self.trace_cfg.get("level", "")).strip().lower()
        if trace_level:
            env["REPLAY_TRACE_LEVEL"] = trace_level
            env["REPLAY_TRACE_SAMPLES"] = str(self.trace_cfg.get("samples", 1))
            env["REPLAY_TRACE_MAX_CHARS"] = str(self.trace_cfg.get("dump_max_chars", 6000))
            env["REPLAY_TRACE_DIR"] = str(self.task_root(task) / "_trace")
            env["REPLAY_DUMP_DIR"] = str(self.task_root(task) / "_trace")
            env["REPLAY_DUMP_MAX_CHARS"] = str(self.trace_cfg.get("dump_max_chars", 6000))
            if trace_level in {"summary", "full"}:
                env["REPLAY_STAGE_DEBUG"] = "1"
                env["REPLAY_STAGE_DEBUG_SAMPLES"] = str(self.trace_cfg.get("samples", 1))
                env["REPLAY_PROMPT_AUDIT"] = "1"
                env["REPLAY_PROMPT_AUDIT_PRINT"] = str(int(truthy(self.trace_cfg.get("prompt_audit_print", False))))
        env["REPLAY_VLLM_RUNTIME_CONTRACT"] = str(
            int(trace_level == "full" or self.explicit_visual_token_shift_axis)
        )
        env["REPLAY_VLLM_RUNTIME_CONTRACT_LEVEL"] = (
            "full" if trace_level == "full" else "count"
        )
        env["REPLAY_VISUAL_TOKEN_SHIFT_FULL_VALIDATION"] = str(
            int(trace_level == "full")
        )
        allowlist_path = self.dataset_index_allowlists.get(task.dataset)
        if allowlist_path:
            env["DATASET_INDEX_ALLOWLIST_FILE"] = str(allowlist_path)
        if task.mode == "image_text_blankimg":
            blank_asset = Path(self.repo_root) / "assets" / "blank-white-1x1.png"
            if not blank_asset.exists():
                raise FileNotFoundError(f"blank image asset not found: {blank_asset}")
            env["REPLAY_BLANK_IMAGE_POSITIONS"] = "2"
            env["REPLAY_BLANK_IMAGE_PATH"] = str(blank_asset)
        max_num_seqs = self.max_num_seqs_for_task(model, task)
        env["VLMEVAL_EFFECTIVE_INFER_BATCH_SIZE"] = str(
            self.infer_batch_size_for_task(model, task)
        )
        env["MINICPM45_VLLM_TP_SIZE"] = str(model.tp_size)
        env["MINICPM45_VLLM_MAX_NUM_SEQS"] = str(max_num_seqs)
        env["GEMMA3_VLLM_TP_SIZE"] = str(model.tp_size)
        env["GEMMA3_VLLM_MAX_NUM_SEQS"] = str(max_num_seqs)
        env["VLLM_TP_SIZE"] = str(model.tp_size)
        env["VLLM_MAX_NUM_SEQS"] = str(max_num_seqs)
        if self._logicvista_qwen25vl_v0_enabled(model, task):
            env["VLLM_USE_V1"] = "0"
            env["VLLM_MAX_NUM_SEQS"] = env.get("LOGICVISTA_QWEN25VL_MAX_NUM_SEQS", "128")
        if task.dataset == "DynaMath":
            env["DYNAMATH_PROMPT_SCHEMA"] = "legacy_two_keys" if self._is_qwen25vl_model(model) else "short_answer_only"
        else:
            env.pop("DYNAMATH_PROMPT_SCHEMA", None)
        if model.max_model_len is not None:
            env["MINICPM45_VLLM_MAX_MODEL_LEN"] = str(model.max_model_len)
            env["GEMMA3_VLLM_MAX_MODEL_LEN"] = str(model.max_model_len)
            env["VLLM_MAX_MODEL_LEN"] = str(model.max_model_len)
        return env

    def run_worker(self, model: ModelSpec, worker_idx: int, gpu_ids: list[str], task_queue: queue.Queue[Task]) -> None:
        initial_qsize = task_queue.qsize()
        claimed = 0
        state_lock = threading.Lock()
        state = {"task_tag": "idle", "phase": "idle"}
        stop_event = threading.Event()
        monitor_thread: threading.Thread | None = None
        if self.worker_monitor_enabled:
            monitor_thread = threading.Thread(
                target=self.monitor_worker,
                args=(model, worker_idx, gpu_ids, self.worker_log_path(model, worker_idx), state_lock, state, stop_event),
                daemon=True,
            )
            monitor_thread.start()
        print(
            f"[WORKER][START] node_rank={self.args.node_rank} model={model.key} "
            f"slot={worker_idx} gpus={','.join(gpu_ids)} queue={initial_qsize}",
            flush=True,
        )
        while True:
            try:
                task = task_queue.get_nowait()
            except queue.Empty:
                break
            claimed += 1
            with state_lock:
                state["task_tag"] = task.tag
                state["phase"] = "running"
            try:
                self.run_single_task(model, task, gpu_ids)
            except Exception as exc:
                self.record_failure(task, f"{type(exc).__name__}: {exc}")
                print(f"[WORKER][FAIL] {task.tag}: {type(exc).__name__}: {exc}", flush=True)
                with state_lock:
                    state["phase"] = "failed"
            finally:
                task_queue.task_done()
                with state_lock:
                    state["task_tag"] = "idle"
                    state["phase"] = "idle"
        stop_event.set()
        if monitor_thread is not None:
            monitor_thread.join(timeout=35)
        print(
            f"[WORKER][DONE] node_rank={self.args.node_rank} model={model.key} slot={worker_idx} completed={claimed}",
            flush=True,
        )

    def task_root(self, task: Task) -> Path:
        dataset_key = task.dataset.replace("/", "_")
        root = self.results_root / task.policy_key / task.mode
        if self.explicit_transform_axis:
            root = root / task.transform
        if self.explicit_visual_token_shift_axis:
            root = root / task.visual_token_shift
        return root / task.model_key / dataset_key

    def model_output_root(self, task: Task, model: ModelSpec) -> Path:
        return self.task_root(task) / model.registry_name

    def prediction_dir(self, task: Task) -> Path:
        return self.task_root(task) / "predictions"

    def eval_output_dir(self, task: Task) -> Path:
        return self.task_root(task) / "eval"

    def prediction_manifest_path(self, task: Task) -> Path:
        return self.prediction_dir(task) / "manifest.json"

    def eval_manifest_path(self, task: Task) -> Path:
        return self.eval_output_dir(task) / "manifest.json"

    def log_root(self, task: Task) -> Path:
        return self.task_root(task) / "_logs"

    def usage_log_path(self, task: Task, model: ModelSpec) -> Path:
        return self.log_root(task) / "usage" / f"{model.registry_name}_{task.dataset}.jsonl"

    def record_failure(self, task: Task, reason: str) -> None:
        message = f"{task.tag}: {reason}"
        with self.failure_lock:
            self.failures.append(message)

    def worker_log_path(self, model: ModelSpec, worker_idx: int) -> Path:
        return self.results_root / "_logs" / "worker_status" / f"node{self.args.node_rank}_{model.key}_slot{worker_idx}.log"

    def infer_file_path(self, task: Task, model: ModelSpec) -> Path | None:
        pred_dir = self.prediction_dir(task)
        for suffix in ("xlsx", "tsv"):
            candidate = pred_dir / f"{model.registry_name}_{task.dataset}.{suffix}"
            if candidate.exists():
                return candidate
        return None

    def infer_artifacts(self, task: Task, model: ModelSpec) -> list[Path]:
        out: list[Path] = []
        pred_dir = self.prediction_dir(task)
        if pred_dir.exists():
            out.append(pred_dir)
        native_dir = self.model_output_root(task, model)
        if native_dir.exists():
            out.append(native_dir)
        return out

    def infer_primary_paths(self, task: Task, model: ModelSpec) -> set[Path]:
        pred_dir = self.prediction_dir(task)
        return {
            pred_dir / f"{model.registry_name}_{task.dataset}.xlsx",
            pred_dir / f"{model.registry_name}_{task.dataset}.tsv",
        }

    def acc_marker_paths(self, task: Task, model: ModelSpec) -> list[Path]:
        eval_dir = self.eval_output_dir(task)
        paths: list[Path] = []
        for pattern in (
            f"*_{task.dataset}*_acc.csv",
            f"*_{task.dataset}*_score.csv",
            f"*_{task.dataset}*_score.json",
        ):
            paths.extend(eval_dir.glob(pattern))
        return sorted(set(paths))

    def acc_complete(self, task: Task, model: ModelSpec) -> bool:
        for path in self.acc_marker_paths(task, model):
            try:
                if path.is_file() and path.stat().st_size > 0:
                    return True
            except OSError:
                continue
        return False

    def eval_artifacts(self, task: Task, model: ModelSpec) -> list[Path]:
        eval_dir = self.eval_output_dir(task)
        return [eval_dir] if eval_dir.exists() else []

    def cleanup_all_artifacts(self, task: Task, model: ModelSpec) -> None:
        for path in self.infer_artifacts(task, model) + self.eval_artifacts(task, model):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)

    def cleanup_eval_artifacts(self, task: Task, model: ModelSpec) -> None:
        for path in self.eval_artifacts(task, model):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)

    def get_expected_count(self, model: ModelSpec, env: dict[str, str], dataset: str) -> int:
        cache_key = (self.env_profiles[model.env_profile].python, dataset)
        if cache_key in self.expected_count_cache:
            return self.expected_count_cache[cache_key]
        code = """
import contextlib
import io
import json
import os
import sys

import pandas as pd


def _normalize_resume_index(raw_idx):
    if raw_idx is None:
        return None
    try:
        if pd.isna(raw_idx):
            return None
    except Exception:
        pass
    if hasattr(raw_idx, "item"):
        try:
            raw_idx = raw_idx.item()
        except Exception:
            pass
    if isinstance(raw_idx, str):
        stripped = raw_idx.strip()
        if stripped == "":
            return stripped
        try:
            return int(stripped)
        except Exception:
            try:
                maybe_float = float(stripped)
            except Exception:
                return stripped
            if maybe_float.is_integer():
                return int(maybe_float)
            return stripped
    if isinstance(raw_idx, float) and raw_idx.is_integer():
        return int(raw_idx)
    return raw_idx


def _load_dataset_index_allowlist():
    raw = os.environ.get("DATASET_INDEX_ALLOWLIST_FILE", "").strip()
    if not raw:
        return None
    if not os.path.exists(raw):
        raise FileNotFoundError(f"DATASET_INDEX_ALLOWLIST_FILE not found: {raw}")
    if raw.lower().endswith(".json"):
        with open(raw, "r", encoding="utf-8") as f:
            payload = json.load(f)
        values = payload.get("indices", []) if isinstance(payload, dict) else payload
    else:
        with open(raw, "r", encoding="utf-8") as f:
            values = [line.strip() for line in f.read().splitlines() if line.strip()]
    if not isinstance(values, list):
        raise ValueError(f"Invalid allowlist payload in {raw}")
    out = set()
    for value in values:
        normalized = _normalize_resume_index(value)
        if normalized is not None:
            out.add(normalized)
    return out


def _filter_data_by_allowlist(data, allowlist):
    if allowlist is None:
        return data
    normalized = data["index"].map(_normalize_resume_index)
    return data[normalized.isin(allowlist)]


name = sys.argv[1]
buf = io.StringIO()
dataset = None
err = None
with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
    try:
        from vlmeval.dataset import build_dataset
        dataset = build_dataset(name)
    except Exception as exc:
        err = exc
if dataset is None or err is not None:
    captured = buf.getvalue().strip()
    if captured:
        print(captured, file=sys.stderr)
    if err is not None:
        print(f"[get_expected_count] build_dataset({name}) failed: {err}", file=sys.stderr)
    print(-1)
    raise SystemExit(0)
try:
    data = getattr(dataset, "data", None)
    if data is None:
        print(int(len(dataset)))
    else:
        data = _filter_data_by_allowlist(data, _load_dataset_index_allowlist())
        print(int(len(data)))
except Exception:
    data = getattr(dataset, "data", None)
    print(int(len(data)) if data is not None else -1)
"""
        proc = subprocess.run(
            [self.env_profiles[model.env_profile].python, "-c", code, dataset],
            cwd=self.repo_root,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        combined = "\n".join(part for part in (proc.stdout, proc.stderr) if part).strip()
        count = -1
        for line in proc.stdout.splitlines():
            line = line.strip()
            if line.lstrip("-").isdigit():
                count = int(line)
        if combined and count < 0:
            print(combined, flush=True)
        self.expected_count_cache[cache_key] = count
        return count

    def _read_json(self, path: Path) -> dict[str, Any]:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        return payload if isinstance(payload, dict) else {}

    def _write_json(self, path: Path, payload: dict[str, Any]) -> None:
        ensure_dir(path.parent)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def _task_payload(self, task: Task, model: ModelSpec, expected: int) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "matrix": self.matrix_name,
            "task": {
                "tag": task.tag,
                "index": task.index,
                "model_key": task.model_key,
                "policy_key": task.policy_key,
                "mode": task.mode,
                "transform": task.transform,
                "visual_token_shift": task.visual_token_shift,
                "dataset": task.dataset,
            },
            "model": {
                "key": model.key,
                "registry_name": model.registry_name,
                "model_path": model.model_path,
            },
            "runtime": {
                "env_profile": model.env_profile,
                "gpus_per_job": model.gpus_per_job,
                "infer_batch_size": self.infer_batch_size_for_task(model, task),
                "max_num_seqs": self.max_num_seqs_for_task(model, task),
                "tp_size": model.tp_size,
            },
            "expected_rows": expected,
            "updated_at": self._timestamp(),
        }

    def visual_token_shift_impl_sha256(self) -> str:
        relative_paths = (
            "run.py",
            "configs/models.yaml",
            "vlmeval/vlm/replay_policy.py",
            "vlmeval/vlm/qwen2_vl/replay_prompt_template.py",
            "vlmeval/vlm/replay_visual_token_shift.py",
            "vlmeval/vlm/replay_vllm_visual_token_models.py",
            "vlmeval/vlm/replay_vllm_visual_token_plugin.py",
            "vlmeval/vlm/qwen2_vl/model.py",
            "vlmeval/vlm/minicpm_v_4_5_replay.py",
            "vlmeval/cli/run_benchmark.py",
        )
        digest = hashlib.sha256()
        for relative_path in relative_paths:
            path = Path(self.repo_root) / relative_path
            digest.update(relative_path.encode("utf-8"))
            digest.update(b"\0")
            digest.update(path.read_bytes())
            digest.update(b"\0")
        digest.update(str(self.matrix_config_path).encode("utf-8"))
        digest.update(b"\0")
        digest.update(self.matrix_config_path.read_bytes())
        digest.update(b"\0")
        digest.update(str(self.model_config_path).encode("utf-8"))
        digest.update(b"\0")
        digest.update(self.model_config_path.read_bytes())
        digest.update(b"\0")
        return digest.hexdigest()

    def visual_token_shift_mechanism_sha256(self) -> str:
        """Hash the mechanism code shared by smoke and production matrices."""
        relative_paths = (
            "run.py",
            "configs/models.yaml",
            "plugins/vllm_visual_token_shift/pyproject.toml",
            "scripts/install_vllm_visual_token_shift_plugin.py",
            "vlmeval/vlm/replay_policy.py",
            "vlmeval/vlm/qwen2_vl/replay_prompt_template.py",
            "vlmeval/vlm/replay_visual_token_shift.py",
            "vlmeval/vlm/replay_vllm_visual_token_models.py",
            "vlmeval/vlm/replay_vllm_visual_token_plugin.py",
            "vlmeval/vlm/qwen2_vl/model.py",
            "vlmeval/vlm/minicpm_v_4_5_replay.py",
            "vlmeval/cli/run_benchmark.py",
        )
        digest = hashlib.sha256()
        for relative_path in relative_paths:
            path = Path(self.repo_root) / relative_path
            digest.update(relative_path.encode("utf-8"))
            digest.update(b"\0")
            digest.update(path.read_bytes())
            digest.update(b"\0")
        return digest.hexdigest()

    def visual_token_shift_validation_harness_sha256(self) -> str:
        """Hash the real-smoke contract that is allowed to issue certificates."""
        relative_paths = (
            "configs/index_allowlists/token_roll_mathvision_first3_20260720.txt",
            "configs/matrix_three_model_iqiq_visual_token_roll_vllm_smoke_20260720.yaml",
            "scripts/run_vllm_visual_token_shift_real_smokes_20260720.sh",
            "scripts/validate_vllm_visual_token_shift_real_dump_20260720.py",
        )
        digest = hashlib.sha256()
        for relative_path in relative_paths:
            path = Path(self.repo_root) / relative_path
            digest.update(relative_path.encode("utf-8"))
            digest.update(b"\0")
            digest.update(path.read_bytes())
            digest.update(b"\0")
        return digest.hexdigest()

    def _allowlist_identity(self, task: Task) -> dict[str, Any]:
        raw_path = self.dataset_index_allowlists.get(task.dataset)
        if not raw_path:
            return {"path": None, "sha256": None}
        path = Path(raw_path).resolve()
        return {
            "path": str(path),
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        }

    @staticmethod
    def _stream_sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            while chunk := handle.read(8 * 1024 * 1024):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _sampled_sha256(path: Path) -> str:
        size = path.stat().st_size
        sample_size = 1024 * 1024
        if size <= 3 * sample_size:
            return BenchmarkRunner._stream_sha256(path)
        digest = hashlib.sha256()
        digest.update(str(size).encode("ascii"))
        with path.open("rb") as handle:
            for offset in (0, max(0, size // 2 - sample_size // 2), size - sample_size):
                handle.seek(offset)
                digest.update(str(offset).encode("ascii"))
                digest.update(handle.read(sample_size))
        return digest.hexdigest()

    def _file_content_identity(self, path: Path, *, sampled: bool = False) -> dict[str, Any]:
        resolved = path.resolve()
        stat = resolved.stat()
        signature = (stat.st_size, stat.st_mtime_ns)
        cache_key = f"{'sampled' if sampled else 'full'}:{resolved}"
        with self.content_identity_lock:
            cached = self.content_identity_cache.get(cache_key)
            if cached is not None and cached[0] == signature:
                return dict(cached[1])
            payload = {
                "path": str(resolved),
                "size": stat.st_size,
                "mtime_ns": stat.st_mtime_ns,
                "digest_kind": "sampled_head_middle_tail_1m" if sampled else "full_sha256",
                "sha256": self._sampled_sha256(resolved) if sampled else self._stream_sha256(resolved),
            }
            self.content_identity_cache[cache_key] = (signature, payload)
            return dict(payload)

    def _source_tree_identity(self, root: Path) -> dict[str, Any]:
        resolved = root.resolve()
        paths = sorted(path for path in resolved.rglob("*.py") if path.is_file())
        signature = tuple(
            (str(path.relative_to(resolved)), path.stat().st_size, path.stat().st_mtime_ns)
            for path in paths
        )
        cache_key = f"python-tree:{resolved}"
        with self.content_identity_lock:
            cached = self.content_identity_cache.get(cache_key)
            if cached is not None and cached[0] == signature:
                return dict(cached[1])
            digest = hashlib.sha256()
            for path in paths:
                digest.update(str(path.relative_to(resolved)).encode("utf-8"))
                digest.update(b"\0")
                digest.update(path.read_bytes())
                digest.update(b"\0")
            payload = {
                "root": str(resolved),
                "python_file_count": len(paths),
                "python_source_sha256": digest.hexdigest(),
            }
            self.content_identity_cache[cache_key] = (signature, payload)
            return dict(payload)

    def _dataset_identity(self, task: Task, env: dict[str, str]) -> dict[str, Any]:
        data_root = Path(env.get("LMUData", "")).expanduser()
        candidates = [
            data_root / f"{task.dataset}{suffix}"
            for suffix in (".tsv", ".json", ".jsonl", ".xlsx")
        ]
        paths = [path for path in candidates if path.is_file()]
        if not paths:
            raise FileNotFoundError(
                f"cannot fingerprint dataset artifact for {task.dataset} under {data_root}"
            )
        return {
            "dataset": task.dataset,
            "artifacts": [self._file_content_identity(path) for path in paths],
            "media_binding": "dataset artifact includes the benchmark image payload or path field",
        }

    def _model_identity(self, model: ModelSpec) -> dict[str, Any]:
        path = Path(model.model_path).expanduser()
        payload: dict[str, Any] = {
            "configured_path": model.model_path,
            "exists": path.exists(),
            "resolved_path": str(path.resolve()) if path.exists() else None,
            "files": [],
        }
        if not path.is_dir():
            return payload
        files = []
        for pattern in ("*.json", "*.py", "*.txt", "*.model", "*.safetensors", "*.bin"):
            for candidate in path.glob(pattern):
                if not candidate.is_file():
                    continue
                stat = candidate.stat()
                record = {"name": candidate.name}
                if candidate.suffix in {".safetensors", ".bin"} and stat.st_size > 8 * 1024 * 1024:
                    record.update(self._file_content_identity(candidate, sampled=True))
                else:
                    record.update(self._file_content_identity(candidate))
                files.append(record)
        payload["files"] = sorted(files, key=lambda item: item["name"])
        return payload

    def _runtime_identity(self, model: ModelSpec, env: dict[str, str]) -> dict[str, Any]:
        profile = self.env_profiles[model.env_profile]
        script = """
import importlib.util
import json
import platform
import sys
from importlib.metadata import PackageNotFoundError, version

names = ["vllm", "transformers", "torch"]
versions = {}
source_roots = {}
for name in names:
    try:
        versions[name] = version(name)
    except PackageNotFoundError:
        versions[name] = None
    spec = importlib.util.find_spec(name)
    roots = list(spec.submodule_search_locations or []) if spec is not None else []
    source_roots[name] = roots[0] if roots else None
print(json.dumps({
    "python": sys.executable,
    "python_version": platform.python_version(),
    "packages": versions,
    "source_roots": source_roots,
}, sort_keys=True))
"""
        proc = subprocess.run(
            [profile.python, "-c", script],
            cwd=self.repo_root,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if proc.returncode != 0:
            raise RuntimeError(
                f"failed to fingerprint runtime {model.env_profile}: {proc.stderr.strip()}"
            )
        payload = json.loads(proc.stdout.strip())
        payload["source_identity"] = {
            name: self._source_tree_identity(Path(root))
            for name, root in payload.pop("source_roots").items()
            if name in {"vllm", "transformers"} and root
        }
        return payload

    def inference_provenance(
        self,
        task: Task,
        model: ModelSpec,
        env: dict[str, str],
    ) -> dict[str, Any]:
        excluded_env = {
            "CUDA_VISIBLE_DEVICES",
            "REPLAY_INFERENCE_FINGERPRINT",
            "REPLAY_VISUAL_TOKEN_SHIFT_RUN_ID",
        }
        prefixes = (
            "REPLAY_",
            "VLLM_",
            "QWEN2VL_",
            "MINICPM45_",
            "LOGICVISTA_",
            "VLMEVAL_",
        )
        semantic_env = {
            key: value
            for key, value in sorted(env.items())
            if key not in excluded_env
            and (key == "PYTHONPATH" or key == "DATASET_INDEX_ALLOWLIST_FILE" or key.startswith(prefixes))
        }
        model_identity = self._model_identity(model)
        runtime_identity = self._runtime_identity(model, env)
        components = {
            "schema_version": 2,
            "implementation_sha256": self.visual_token_shift_impl_sha256(),
            "mechanism_implementation_sha256": self.visual_token_shift_mechanism_sha256(),
            "validation_harness_sha256": (
                self.visual_token_shift_validation_harness_sha256()
            ),
            "repository_source_identity": self._source_tree_identity(
                Path(self.repo_root) / "vlmeval"
            ),
            "allowlist": self._allowlist_identity(task),
            "dataset_identity": self._dataset_identity(task, env),
            "model_identity": model_identity,
            "runtime_identity": runtime_identity,
            "semantic_env": semantic_env,
        }
        fingerprint = hashlib.sha256(
            json.dumps(components, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
        return {**components, "fingerprint": fingerprint}

    @staticmethod
    def _smoke_binding_components(
        provenance: dict[str, Any],
        *,
        model_key: str,
        model_family: str,
        mode: str,
        engine_mode: str,
    ) -> dict[str, Any]:
        runtime = provenance.get("runtime_identity") or {}
        runtime_sources = runtime.get("source_identity") or {}
        model = provenance.get("model_identity") or {}
        normalized_files = sorted(
            (
                {
                    key: item.get(key)
                    for key in ("name", "size", "digest_kind", "sha256")
                }
                for item in model.get("files", [])
            ),
            key=lambda item: str(item.get("name")),
        )
        model_content = {
            "configured_path": model.get("configured_path"),
            "resolved_path": model.get("resolved_path"),
            "files": normalized_files,
        }
        model_content_sha256 = hashlib.sha256(
            json.dumps(model_content, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
        return {
            "schema_version": 1,
            "model_key": model_key,
            "model_family": model_family,
            "mode": mode,
            "vllm_engine_mode": engine_mode,
            "mechanism_implementation_sha256": provenance.get(
                "mechanism_implementation_sha256"
            ),
            "validation_harness_sha256": provenance.get(
                "validation_harness_sha256"
            ),
            "repository_python_source_sha256": (
                provenance.get("repository_source_identity") or {}
            ).get("python_source_sha256"),
            "runtime": {
                "python": runtime.get("python"),
                "python_version": runtime.get("python_version"),
                "packages": runtime.get("packages"),
                "vllm_python_source_sha256": (
                    runtime_sources.get("vllm") or {}
                ).get("python_source_sha256"),
                "transformers_python_source_sha256": (
                    runtime_sources.get("transformers") or {}
                ).get("python_source_sha256"),
            },
            "model": {
                "configured_path": model.get("configured_path"),
                "resolved_path": model.get("resolved_path"),
                "content_sha256": model_content_sha256,
            },
        }

    def verify_vllm_visual_token_shift_smoke_certificate(
        self,
        task: Task,
        model: ModelSpec,
        env: dict[str, str],
        inference_provenance: dict[str, Any],
    ) -> dict[str, Any]:
        certificate_path_raw = env.get(
            "REPLAY_VISUAL_TOKEN_SHIFT_SMOKE_CERTIFICATE", ""
        ).strip()
        if not certificate_path_raw:
            raise RuntimeError(
                "lightweight vLLM token shift requires "
                "REPLAY_VISUAL_TOKEN_SHIFT_SMOKE_CERTIFICATE"
            )
        certificate_path = Path(certificate_path_raw).expanduser().resolve()
        certificate_bytes = certificate_path.read_bytes()
        certificate = json.loads(certificate_bytes)
        smoke_certificate = certificate.get("smoke_certificate") or {}
        if (
            certificate.get("all_passed") is not True
            or smoke_certificate.get("schema_version") != 1
            or smoke_certificate.get("valid") is not True
        ):
            raise RuntimeError("vLLM token-shift smoke certificate is not valid")
        expected_components = self._smoke_binding_components(
            inference_provenance,
            model_key=model.key,
            model_family=env.get("REPLAY_VLLM_TARGET_FAMILY", ""),
            mode=task.visual_token_shift,
            engine_mode="v0" if env.get("VLLM_USE_V1") == "0" else "v1",
        )
        expected_fingerprint = hashlib.sha256(
            json.dumps(
                expected_components, ensure_ascii=False, sort_keys=True
            ).encode("utf-8")
        ).hexdigest()
        matches = [
            entry
            for entry in smoke_certificate.get("entries", [])
            if entry.get("binding_fingerprint") == expected_fingerprint
            and entry.get("binding") == expected_components
        ]
        if len(matches) != 1:
            raise RuntimeError(
                "no unique full-smoke certificate matches the current "
                f"code/runtime/model/mode identity: {expected_fingerprint}"
            )
        return {
            "schema_version": 1,
            "path": str(certificate_path),
            "file_sha256": hashlib.sha256(certificate_bytes).hexdigest(),
            "binding_fingerprint": expected_fingerprint,
        }

    def prediction_file_valid(self, pred_file: Path | None, expected: int) -> bool:
        if expected < 0 or pred_file is None or not pred_file.exists():
            return False
        try:
            rows = load_tabular_rows(pred_file)
        except Exception:
            return False
        if len(rows) != expected:
            return False
        candidate_cols = [col for col in ("prediction", "description", "detailed_prediction") if rows and col in rows[0]]
        if not candidate_cols:
            return True
        for row in rows:
            values = [row.get(col) for col in candidate_cols]
            if all(is_blank(value) for value in values):
                return False
            desc = str(row.get("description", "")).strip()
            if desc.startswith("[FAILED_INFER]") or "Failed to obtain answer via API." in desc:
                return False
        return True

    def write_prediction_manifest(
        self,
        task: Task,
        model: ModelSpec,
        expected: int,
        pred_file: Path,
        mechanism_evidence: dict[str, Any] | None = None,
        runtime_contract_run_id: str | None = None,
        inference_provenance: dict[str, Any] | None = None,
    ) -> None:
        payload = self._task_payload(task, model, expected)
        payload.update(
            {
                "artifact_type": "prediction",
                "status": "complete",
                "prediction_file": str(pred_file),
                "prediction_dir": str(self.prediction_dir(task)),
                "mechanism_evidence": mechanism_evidence,
                "runtime_contract_run_id": runtime_contract_run_id,
                "inference_provenance": inference_provenance,
            }
        )
        self._write_json(self.prediction_manifest_path(task), payload)

    def write_eval_manifest(self, task: Task, model: ModelSpec, expected: int, pred_file: Path, rc: int) -> None:
        score_files = [str(path) for path in self.acc_marker_paths(task, model) if path.is_file()]
        payload = self._task_payload(task, model, expected)
        payload.update(
            {
                "artifact_type": "eval",
                "status": "complete" if rc == 0 and score_files else "failed",
                "prediction_file": str(pred_file),
                "eval_dir": str(self.eval_output_dir(task)),
                "judge": str(self.evaluation_cfg.get("judge", "gpt-4o-mini")),
                "score_files": score_files,
                "returncode": rc,
            }
        )
        self._write_json(self.eval_manifest_path(task), payload)

    def infer_complete(
        self,
        task: Task,
        model: ModelSpec,
        expected: int,
        env: dict[str, str],
    ) -> bool:
        if expected < 0:
            return False
        manifest = self._read_json(self.prediction_manifest_path(task))
        if manifest.get("status") != "complete" or int(manifest.get("expected_rows", -1)) != expected:
            return False
        current_provenance = self.inference_provenance(task, model, env)
        if (
            (manifest.get("inference_provenance") or {}).get("fingerprint")
            != current_provenance["fingerprint"]
        ):
            return False
        matched_token_roll_runtime = truthy(
            env.get("REPLAY_VLLM_MATCHED_TOKEN_ROLL_RUNTIME", "0")
        )
        if task.visual_token_shift != "none" or matched_token_roll_runtime:
            evidence = manifest.get("mechanism_evidence") or {}
            expected_runtime = {
                "env_profile": model.env_profile,
                "gpus_per_job": model.gpus_per_job,
                "infer_batch_size": self.infer_batch_size_for_task(model, task),
                "max_num_seqs": self.max_num_seqs_for_task(model, task),
                "tp_size": model.tp_size,
            }
            if (
                evidence.get("verified") is not True
                or evidence.get("mode") != task.visual_token_shift
                or (
                    task.visual_token_shift == "none"
                    and evidence.get("control") is not True
                )
                or evidence.get("model_key") != model.key
                or evidence.get("matrix") != self.matrix_name
                or evidence.get("model_path") != model.model_path
                or evidence.get("run_id")
                != manifest.get("runtime_contract_run_id")
                or evidence.get("inference_fingerprint")
                != current_provenance["fingerprint"]
                or manifest.get("runtime") != expected_runtime
            ):
                return False
            if task.visual_token_shift != "none" and not truthy(
                env.get("REPLAY_VISUAL_TOKEN_SHIFT_FULL_VALIDATION", "0")
            ):
                try:
                    current_certificate = (
                        self.verify_vllm_visual_token_shift_smoke_certificate(
                            task, model, env, current_provenance
                        )
                    )
                except Exception:
                    return False
                if evidence.get("smoke_certificate") != current_certificate:
                    return False
        pred_file = self.infer_file_path(task, model)
        return self.prediction_file_valid(pred_file, expected)

    def existing_infer_complete(
        self,
        task: Task,
        model: ModelSpec,
        expected: int,
        env: dict[str, str],
    ) -> bool:
        """Validate a producer matrix's inference without taking ownership of it."""
        if expected < 0:
            return False
        manifest = self._read_json(self.prediction_manifest_path(task))
        if (
            manifest.get("status") != "complete"
            or manifest.get("matrix") != self.inference_producer_matrix
            or int(manifest.get("expected_rows", -1)) != expected
        ):
            return False
        recorded_provenance = manifest.get("inference_provenance") or {}
        recorded_fingerprint = recorded_provenance.get("fingerprint")
        if not recorded_fingerprint:
            return False
        current_provenance = self.inference_provenance(task, model, env)
        immutable_components = (
            "mechanism_implementation_sha256",
            "validation_harness_sha256",
            "repository_source_identity",
            "allowlist",
            "dataset_identity",
            "model_identity",
            "runtime_identity",
        )
        if any(
            recorded_provenance.get(key) != current_provenance.get(key)
            for key in immutable_components
        ):
            return False
        ignored_semantic_env = {"REPLAY_VISUAL_TOKEN_SHIFT_SMOKE_CERTIFICATE"}
        recorded_semantic_env = {
            key: value
            for key, value in (recorded_provenance.get("semantic_env") or {}).items()
            if key not in ignored_semantic_env
        }
        current_semantic_env = {
            key: value
            for key, value in (current_provenance.get("semantic_env") or {}).items()
            if key not in ignored_semantic_env
        }
        if recorded_semantic_env != current_semantic_env:
            return False
        expected_task = {
            "tag": task.tag,
            "index": task.index,
            "model_key": task.model_key,
            "policy_key": task.policy_key,
            "mode": task.mode,
            "transform": task.transform,
            "visual_token_shift": task.visual_token_shift,
            "dataset": task.dataset,
        }
        expected_model = {
            "key": model.key,
            "registry_name": model.registry_name,
            "model_path": model.model_path,
        }
        expected_runtime = {
            "env_profile": model.env_profile,
            "gpus_per_job": model.gpus_per_job,
            "infer_batch_size": self.infer_batch_size_for_task(model, task),
            "max_num_seqs": self.max_num_seqs_for_task(model, task),
            "tp_size": model.tp_size,
        }
        if (
            manifest.get("task") != expected_task
            or manifest.get("model") != expected_model
            or manifest.get("runtime") != expected_runtime
        ):
            return False
        matched_token_roll_runtime = truthy(
            env.get("REPLAY_VLLM_MATCHED_TOKEN_ROLL_RUNTIME", "0")
        )
        if task.visual_token_shift != "none" or matched_token_roll_runtime:
            evidence = manifest.get("mechanism_evidence") or {}
            if (
                evidence.get("verified") is not True
                or evidence.get("mode") != task.visual_token_shift
                or (
                    task.visual_token_shift == "none"
                    and evidence.get("control") is not True
                )
                or evidence.get("model_key") != model.key
                or evidence.get("matrix") != self.inference_producer_matrix
                or evidence.get("model_path") != model.model_path
                or evidence.get("run_id")
                != manifest.get("runtime_contract_run_id")
                or evidence.get("inference_fingerprint") != recorded_fingerprint
            ):
                return False
            if task.visual_token_shift == "none":
                pred_file = self.infer_file_path(task, model)
                return self.prediction_file_valid(pred_file, expected)
            stored_certificate = evidence.get("smoke_certificate") or {}
            certificate_path = str(stored_certificate.get("path", "")).strip()
            if not certificate_path:
                return False
            certificate_env = dict(env)
            certificate_env["REPLAY_VISUAL_TOKEN_SHIFT_SMOKE_CERTIFICATE"] = (
                certificate_path
            )
            try:
                current_certificate = (
                    self.verify_vllm_visual_token_shift_smoke_certificate(
                        task, model, certificate_env, recorded_provenance
                    )
                )
            except Exception:
                return False
            if stored_certificate != current_certificate:
                return False
        pred_file = self.infer_file_path(task, model)
        return self.prediction_file_valid(pred_file, expected)

    def eval_complete(self, task: Task, model: ModelSpec, expected: int) -> bool:
        manifest = self._read_json(self.eval_manifest_path(task))
        if manifest.get("status") != "complete" or int(manifest.get("expected_rows", -1)) != expected:
            return False
        return self.acc_complete(task, model)

    def run_subprocess(self, cmd: list[str], env: dict[str, str], log_path: Path) -> int:
        ensure_dir(log_path.parent)
        with log_path.open("w", encoding="utf-8") as log_fh:
            proc = subprocess.run(
                cmd,
                cwd=self.repo_root,
                env=env,
                stdout=log_fh,
                stderr=subprocess.STDOUT,
                text=True,
            )
        return proc.returncode

    def _query_nvidia_smi(self, gpu_ids: list[str]) -> str:
        gpu_arg = ",".join(gpu_ids)
        proc = subprocess.run(
            ["nvidia-smi", "-i", gpu_arg],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            timeout=30,
        )
        return proc.stdout.strip()

    def _write_worker_snapshot(
        self,
        log_path: Path,
        model: ModelSpec,
        worker_idx: int,
        gpu_ids: list[str],
        state_lock: threading.Lock,
        state: dict[str, str],
        reason: str,
    ) -> None:
        ensure_dir(log_path.parent)
        with state_lock:
            task_tag = state.get("task_tag", "idle")
            phase = state.get("phase", "idle")
        try:
            smi_text = self._query_nvidia_smi(gpu_ids)
        except Exception as exc:
            smi_text = f"[nvidia-smi failed] {type(exc).__name__}: {exc}"
        payload = {
            "timestamp": self._timestamp(),
            "node_rank": self.args.node_rank,
            "model": model.key,
            "worker_idx": worker_idx,
            "gpu_ids": gpu_ids,
            "reason": reason,
            "task_tag": task_tag,
            "phase": phase,
        }
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(payload, ensure_ascii=False) + "\n")
            fh.write(smi_text + "\n\n")

    def monitor_worker(
        self,
        model: ModelSpec,
        worker_idx: int,
        gpu_ids: list[str],
        log_path: Path,
        state_lock: threading.Lock,
        state: dict[str, str],
        stop_event: threading.Event,
    ) -> None:
        self._write_worker_snapshot(log_path, model, worker_idx, gpu_ids, state_lock, state, reason="start")
        while not stop_event.wait(self.worker_monitor_interval):
            self._write_worker_snapshot(log_path, model, worker_idx, gpu_ids, state_lock, state, reason="periodic")
        self._write_worker_snapshot(log_path, model, worker_idx, gpu_ids, state_lock, state, reason="stop")

    def run_answer_format(self, task: Task, model: ModelSpec, env: dict[str, str]) -> None:
        if not truthy(self.answer_format_cfg.get("enable", 1)):
            return
        pred_file = self.infer_file_path(task, model)
        if pred_file is None:
            return
        eval_dir = self.eval_output_dir(task)
        ensure_dir(eval_dir)
        report = eval_dir / f"{model.registry_name}_{task.dataset}_answer_format_report.json"
        failures = eval_dir / f"{model.registry_name}_{task.dataset}_answer_format_failures.jsonl"
        log_path = self.log_root(task) / "answer_format" / f"{model.registry_name}_{task.dataset}_{self._timestamp()}.log"
        cmd = [
            self.env_profiles[model.env_profile].python,
            "-m",
            "vlmeval.cli.postprocess_answer_format",
            "--pred-file",
            str(pred_file),
            "--out-json",
            str(report),
            "--out-fail-jsonl",
            str(failures),
            "--response-col",
            str(self.answer_format_cfg.get("response_col", "prediction")),
            "--fallback-col",
            str(self.answer_format_cfg.get("fallback_col", "detailed_prediction")),
            "--require-boxed",
            str(self.answer_format_cfg.get("require_boxed", 0)),
            "--max-fails",
            str(self.answer_format_cfg.get("max_fails", 50)),
        ]
        rc = self.run_subprocess(cmd, env, log_path)
        if rc != 0:
            print(f"[FAIL][FORMAT] {task.tag}: rc={rc} log={log_path}", flush=True)

    def verify_vllm_visual_token_shift_run(
        self,
        task: Task,
        model: ModelSpec,
        env: dict[str, str],
        expected_count: int,
        inference_provenance: dict[str, Any],
    ) -> dict[str, Any]:
        run_id = env.get("REPLAY_VISUAL_TOKEN_SHIFT_RUN_ID", "")
        model_family = env.get("REPLAY_VLLM_TARGET_FAMILY", "")
        inference_fingerprint = inference_provenance["fingerprint"]
        dump_dir = Path(env["REPLAY_VISUAL_TOKEN_SHIFT_DUMP_DIR"])
        expected_ranks = set(range(model.tp_size))
        expected_engine_mode = "v0" if env.get("VLLM_USE_V1") == "0" else "v1"
        full_validation = truthy(
            env.get("REPLAY_VISUAL_TOKEN_SHIFT_FULL_VALIDATION", "0")
        )
        contracts = []
        for path in dump_dir.glob("vllm_runtime_contract.pid*.jsonl"):
            for line in path.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                payload = json.loads(line)
                if payload.get("run_id") == run_id:
                    contracts.append(payload)
        if not contracts:
            raise RuntimeError("missing vLLM runtime request contract")
        contracts.sort(key=lambda item: float(item.get("timestamp", 0.0)))
        if any(
            item.get("phase") != "vllm_runtime_contract"
            or item.get("model_family") != model_family
            or item.get("mode") != task.visual_token_shift
            or item.get("inference_fingerprint") != inference_fingerprint
            or int(item.get("effective_infer_batch_size", -1))
            != self.infer_batch_size_for_task(model, task)
            or int(item.get("effective_max_num_seqs", -1))
            != self.max_num_seqs_for_task(model, task)
            or item.get("actual_vllm_engine_mode") != expected_engine_mode
            or not item.get("actual_vllm_engine_module")
            or not item.get("actual_vllm_engine_qualname")
            or not item.get("actual_vllm_runtime_config")
            for item in contracts
        ):
            raise RuntimeError("invalid vLLM runtime request contract")
        contract_request_count = sum(int(item.get("request_count", -1)) for item in contracts)
        if contract_request_count != expected_count:
            raise RuntimeError(
                f"runtime request coverage mismatch: expected={expected_count} "
                f"contract={contract_request_count}"
            )
        contract_request_counts = [
            int(item.get("request_count", -1)) for item in contracts
        ]
        expected_request_counts = []
        remaining_request_count = expected_count
        effective_batch_size = self.infer_batch_size_for_task(model, task)
        while remaining_request_count > 0:
            call_request_count = min(effective_batch_size, remaining_request_count)
            expected_request_counts.append(call_request_count)
            remaining_request_count -= call_request_count
        if contract_request_counts != expected_request_counts:
            raise RuntimeError(
                "runtime request call partition mismatch: "
                f"expected={expected_request_counts} "
                f"contract={contract_request_counts}"
            )
        runtime_configs = [
            item.get("actual_vllm_runtime_config") or {} for item in contracts
        ]
        if len({json.dumps(item, sort_keys=True) for item in runtime_configs}) != 1:
            raise RuntimeError("concrete vLLM runtime config changed across calls")
        runtime_config = runtime_configs[0]
        if (
            int(runtime_config.get("max_num_seqs") or -1)
            != self.max_num_seqs_for_task(model, task)
            or int(runtime_config.get("tensor_parallel_size") or -1) != model.tp_size
            or runtime_config.get("enable_chunked_prefill") is not False
            or runtime_config.get("disable_chunked_mm_input") is not True
            or runtime_config.get("enable_prefix_caching") is not False
            or int(runtime_config.get("max_model_len") or 0) <= 0
            or int(runtime_config.get("max_num_batched_tokens") or 0) <= 0
            or not isinstance(runtime_config.get("limit_mm_per_prompt"), dict)
        ):
            raise RuntimeError("invalid concrete vLLM runtime config")
        request_metadata_by_call = [
            list(contract.get("request_metadata", [])) for contract in contracts
        ]
        request_metadata = [
            metadata for call_metadata in request_metadata_by_call
            for metadata in call_metadata
        ]
        if full_validation and model_family == "minicpm_o_4_5":
            if any(
                len(call_metadata) != call_request_count
                or [
                    int(metadata.get("request_index", -1))
                    for metadata in call_metadata
                ]
                != list(range(call_request_count))
                for call_metadata, call_request_count in zip(
                    request_metadata_by_call,
                    contract_request_counts,
                )
            ) or len(request_metadata) != contract_request_count or any(
                metadata.get("image_count") != 2
                or len(metadata.get("image_sha256") or []) != 2
                or len(metadata.get("minicpm_num_slices") or []) != 2
                or metadata["image_sha256"][0] != metadata["image_sha256"][1]
                or metadata["minicpm_num_slices"][0]
                != metadata["minicpm_num_slices"][1]
                for metadata in request_metadata
            ):
                raise RuntimeError("invalid MiniCPM driver slice metadata")
            if any(
                len(
                    {
                        tuple(
                            int(value)
                            for value in metadata.get(
                                "minicpm_num_slices",
                                [],
                            )
                        )
                        for metadata in call_metadata
                    }
                )
                != len(call_metadata)
                for call_metadata in request_metadata_by_call
                if len(call_metadata) > 1
            ):
                raise RuntimeError(
                    "MiniCPM sentinel slices are not request-distinguishing "
                    "within a driver call"
                )
        handshakes = []
        for path in dump_dir.glob(f"worker_handshake.{model_family}.pid*.json"):
            payload = self._read_json(path)
            if (
                payload.get("run_id") == run_id
                and payload.get("model_family") == model_family
                and payload.get("mode") == task.visual_token_shift
            ):
                handshakes.append(payload)
        handshake_ranks = {int(item["rank"]) for item in handshakes}
        if handshake_ranks != expected_ranks or len(handshakes) != len(expected_ranks):
            raise RuntimeError(
                f"worker handshake ranks mismatch: expected={sorted(expected_ranks)} "
                f"observed={sorted(handshake_ranks)} records={len(handshakes)}"
            )
        if any(
            item.get("phase") != "worker_model_initialized"
            or item.get("backend") != "vllm"
            or item.get("target_family") != model_family
            or item.get("inference_fingerprint") != inference_fingerprint
            or int(item.get("scheduler_max_num_seqs", -1))
            != self.max_num_seqs_for_task(model, task)
            or int(item.get("scheduler_max_num_batched_tokens", -1))
            != int(runtime_config.get("max_num_batched_tokens", -2))
            or item.get("scheduler_enable_chunked_prefill") is not False
            or item.get("scheduler_disable_chunked_mm_input") is not True
            or item.get("cache_enable_prefix_caching") is not False
            or item.get("vllm_engine_mode") != expected_engine_mode
            or not item.get("model_class")
            or not item.get("model_name")
            for item in handshakes
        ):
            raise RuntimeError("invalid vLLM worker handshake metadata")

        expected_stage = (
            "legacy_v0_post_projector_pre_llm"
            if model_family == "qwen2_5_vl" and env.get("VLLM_USE_V1") == "0"
            else "post_gather_pre_llm_get_input_embeddings"
            if model_family == "qwen2_5_vl"
            else "post_gather_pre_llm_embed_input_ids"
        )
        if task.visual_token_shift == "none":
            none_shift_records = []
            for path in dump_dir.glob("visual_token_shift.vllm.pid*.jsonl"):
                for line in path.read_text(encoding="utf-8").splitlines():
                    if line.strip():
                        payload = json.loads(line)
                        if payload.get("run_id") == run_id:
                            none_shift_records.append(payload)
            none_call_records = []
            for path in dump_dir.glob("visual_token_shift_calls.vllm.pid*.jsonl"):
                for line in path.read_text(encoding="utf-8").splitlines():
                    if line.strip():
                        payload = json.loads(line)
                        if payload.get("run_id") == run_id:
                            none_call_records.append(payload)
            if none_shift_records or none_call_records:
                raise RuntimeError(
                    "matched none control unexpectedly shifted visual tokens"
                )
            return {
                "verified": True,
                "control": True,
                "backend": "vllm",
                "mode": "none",
                "model_key": model.key,
                "model_family": model_family,
                "matrix": self.matrix_name,
                "model_path": model.model_path,
                "implementation_sha256": inference_provenance["implementation_sha256"],
                "inference_fingerprint": inference_fingerprint,
                "run_id": run_id,
                "expected_tp_ranks": sorted(expected_ranks),
                "worker_handshake_count": len(handshakes),
                "shift_record_count": 0,
                "runtime_contract_count": len(contracts),
                "validated_request_count": contract_request_count,
                "actual_vllm_engine_mode": contracts[0]["actual_vllm_engine_mode"],
                "actual_vllm_engine_module": contracts[0]["actual_vllm_engine_module"],
                "actual_vllm_engine_qualname": contracts[0]["actual_vllm_engine_qualname"],
                "actual_vllm_runtime_config": runtime_config,
            }
        expected_shift = {"noop_vllm": 0, "roll_right_1": 1, "roll_left_1": -1}.get(
            task.visual_token_shift
        )
        if expected_shift is None:
            raise RuntimeError(f"unsupported audited visual-token shift {task.visual_token_shift}")

        def minicpm_placeholder_contract_exact(item: dict[str, Any]) -> bool:
            query_num = int(
                item.get("minicpm_query_num")
                or (item.get("model_metadata") or {}).get("query_num")
                or 0
            )
            item_counts = [int(value) for value in item.get("item_token_counts") or []]
            span_groups = item.get("item_span_groups") or []
            slice_counts = [
                int(value)
                for value in item.get("minicpm_placeholder_slice_counts") or []
            ]
            return bool(
                query_num == 64
                and item.get("minicpm_placeholder_token_contract_exact") is True
                and len(item_counts) == len(span_groups) == len(slice_counts)
                and slice_counts == [len(group) for group in span_groups]
                and all(
                    all(int(span_length) == query_num for span_length in group)
                    for group in span_groups
                )
                and all(
                    item_count == query_num * slice_count
                    for item_count, slice_count in zip(item_counts, slice_counts)
                )
            )

        records = []
        for path in dump_dir.glob("visual_token_shift.vllm.pid*.jsonl"):
            for line in path.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                payload = json.loads(line)
                if payload.get("run_id") == run_id:
                    records.append(payload)
        record_ranks = {int(item["rank"]) for item in records}
        if record_ranks != expected_ranks:
            raise RuntimeError(
                f"shift record ranks mismatch: expected={sorted(expected_ranks)} "
                f"observed={sorted(record_ranks)}"
            )
        raw_call_indices_by_rank: dict[int, list[int]] = {}
        for rank in expected_ranks:
            rank_records = sorted(
                (item for item in records if int(item["rank"]) == rank),
                key=lambda item: int(item.get("call_index", -1)),
            )
            raw_call_indices = [
                int(item.get("call_index", -1)) for item in rank_records
            ]
            if not rank_records or raw_call_indices != list(
                range(1, len(rank_records) + 1)
            ):
                raise RuntimeError(
                    "raw prefill shifts must be a non-empty contiguous call prefix "
                    f"on rank {rank}, got {raw_call_indices}"
                )
            raw_call_indices_by_rank[rank] = raw_call_indices
            for record in rank_records:
                if (
                    record.get("mode") != task.visual_token_shift
                    or record.get("model_family") != model_family
                    or record.get("phase") != "visual_token_shift"
                    or record.get("backend") != "vllm"
                    or record.get("stage") != expected_stage
                    or record.get("real_request") is not True
                    or record.get("recording_armed") is not True
                    or record.get("inference_fingerprint") != inference_fingerprint
                    or record.get("strict") is not True
                    or record.get("full_tensor_validation") is not full_validation
                    or record.get("validation_level")
                    != ("full" if full_validation else "lightweight")
                ):
                    raise RuntimeError(
                        f"invalid vLLM visual-token shift record on rank {rank}"
                    )
                if full_validation:
                    if (
                        record.get("input_ids_unchanged_exact") is not True
                        or record.get("is_multimodal_unchanged_exact") is not True
                        or record.get("item_token_coverage_exact") is not True
                        or record.get("item_span_grouping_exact") is not True
                        or record.get("final_mm_scatter_exact") is not True
                        or record.get("output_shape_matches_input_ids") is not True
                    ):
                        raise RuntimeError(
                            f"invalid full tensor validation on rank {rank}"
                        )
                    if model_family == "qwen2_5_vl" and record.get(
                        "item_span_lengths_exact"
                    ) is not True:
                        raise RuntimeError(
                            f"invalid Qwen placeholder spans on rank {rank}"
                        )
                    if model_family == "qwen2_5_vl" and record.get(
                        "qwen_grid_token_count_exact"
                    ) is not True:
                        raise RuntimeError(
                            f"invalid Qwen grid token count on rank {rank}"
                        )
                    if model_family == "minicpm_o_4_5" and not (
                        minicpm_placeholder_contract_exact(record)
                    ):
                        raise RuntimeError(
                            f"invalid MiniCPM slice contract on rank {rank}"
                        )
                elif record.get("output_shape_matches_input_ids") is not True:
                    raise RuntimeError(
                        f"invalid lightweight output shape on rank {rank}"
                    )
                if not full_validation and record.get("audit", {}).get(
                    "pair_equality_required"
                ) is not False:
                    raise RuntimeError(
                        "lightweight audit unexpectedly scans pair values on "
                        f"rank {rank}"
                    )
                pairs = record.get("audit", {}).get("pair_records", [])
                if not pairs:
                    raise RuntimeError(f"invalid IQIQ pair audit on rank {rank}")
                for pair in pairs:
                    token_count = int(pair.get("shape", [0])[0])
                    canonical_source = [
                        int((output_index - expected_shift) % token_count)
                        for output_index in range(token_count)
                    ]
                    if (
                        int(pair.get("shift", 99)) != expected_shift
                        or pair.get("source_index_for_output") != canonical_source
                        or pair.get("i2_out_of_place") is not True
                    ):
                        raise RuntimeError(
                            "invalid canonical token roll on rank "
                            f"{rank} pair {pair.get('pair_index')}"
                        )
                    if full_validation and (
                        pair.get("i1_i2_equal_exact") is not True
                        or pair.get("i1_unchanged_exact") is not True
                        or pair.get("i2_source_unchanged_exact") is not True
                        or pair.get("i2_roll_exact") is not True
                        or float(pair.get("max_abs_error", float("inf"))) != 0.0
                    ):
                        raise RuntimeError(
                            "invalid exact tensor roll on rank "
                            f"{rank} pair {pair.get('pair_index')}"
                        )
            if full_validation and model_family == "minicpm_o_4_5":
                raw_pair_counts = [
                    len((record.get("audit") or {}).get("pair_records", []))
                    for record in rank_records
                ]
                if not contiguous_partition_refines(
                    raw_pair_counts,
                    contract_request_counts,
                ):
                    raise RuntimeError(
                        "MiniCPM worker prefill partition is not a contiguous "
                        "refinement of driver calls: "
                        f"worker={raw_pair_counts} driver={contract_request_counts}"
                    )
                driver_slice_pairs = [
                    tuple(
                        int(value)
                        for value in metadata.get("minicpm_num_slices", [])
                    )
                    for metadata in request_metadata
                ]
                worker_slice_pairs = []
                for record in rank_records:
                    pair_records = (record.get("audit") or {}).get(
                        "pair_records",
                        [],
                    )
                    slice_counts = record.get(
                        "minicpm_placeholder_slice_counts"
                    ) or []
                    if len(slice_counts) != 2 * len(pair_records):
                        raise RuntimeError(
                            "MiniCPM worker slice counts do not cover IQIQ pairs"
                        )
                    worker_slice_pairs.extend(
                        tuple(
                            int(value)
                            for value in slice_counts[
                                2 * pair_index : 2 * pair_index + 2
                            ]
                        )
                        for pair_index in range(len(pair_records))
                    )
                if worker_slice_pairs != driver_slice_pairs:
                    raise RuntimeError(
                        "MiniCPM driver and worker slice order differ: "
                        f"driver={driver_slice_pairs} worker={worker_slice_pairs}"
                    )

        call_records = []
        for path in dump_dir.glob("visual_token_shift_calls.vllm.pid*.jsonl"):
            for line in path.read_text(encoding="utf-8").splitlines():
                if line.strip():
                    payload = json.loads(line)
                    if payload.get("run_id") == run_id:
                        call_records.append(payload)
        call_counts_by_rank = {}
        pair_counts_by_rank: dict[int, list[int]] = {}
        for rank in expected_ranks:
            rank_calls = sorted(
                (item for item in call_records if int(item.get("rank", -1)) == rank),
                key=lambda item: int(item.get("call_index", -1)),
            )
            call_indices = [int(item.get("call_index", -1)) for item in rank_calls]
            if call_indices != list(range(1, len(rank_calls) + 1)) or not rank_calls:
                raise RuntimeError(
                    f"non-contiguous or empty prefill call audit on rank {rank}: {call_indices}"
                )
            raw_call_indices = raw_call_indices_by_rank[rank]
            if raw_call_indices != call_indices[: len(raw_call_indices)]:
                raise RuntimeError(
                    "raw prefill calls are not an audited-call prefix on rank "
                    f"{rank}: raw={raw_call_indices} audited={call_indices}"
                )
            for item in rank_calls:
                if (
                    item.get("mode") != task.visual_token_shift
                    or item.get("model_family") != model_family
                    or item.get("real_request") is not True
                    or item.get("recording_armed") is not True
                    or item.get("inference_fingerprint") != inference_fingerprint
                    or item.get("full_tensor_validation") is not full_validation
                    or item.get("validation_level")
                    != ("full" if full_validation else "lightweight")
                    or item.get("pair_structure_validated") is not True
                ):
                    raise RuntimeError(
                        f"invalid all-prefill call audit on rank {rank} call {item.get('call_index')}"
                    )
                if full_validation:
                    if (
                        item.get("all_iqiq_pairs_equal_exact") is not True
                        or item.get("input_ids_unchanged_exact") is not True
                        or item.get("is_multimodal_unchanged_exact") is not True
                        or item.get("final_mm_scatter_exact") is not True
                        or item.get("item_token_coverage_exact") is not True
                        or item.get("item_span_grouping_exact") is not True
                        or item.get("output_shape_matches_input_ids") is not True
                        or (
                            model_family == "qwen2_5_vl"
                            and (
                                item.get("item_span_lengths_exact") is not True
                                or item.get("qwen_grid_token_count_exact") is not True
                            )
                        )
                        or (
                            model_family == "minicpm_o_4_5"
                            and not minicpm_placeholder_contract_exact(item)
                        )
                    ):
                        raise RuntimeError(
                            f"invalid full prefill audit on rank {rank} call {item.get('call_index')}"
                        )
                elif (
                    item.get("output_shape_matches_input_ids") is not True
                    or item.get("all_iqiq_pairs_equal_exact") is not None
                ):
                    raise RuntimeError(
                        f"invalid lightweight prefill audit on rank {rank} call {item.get('call_index')}"
                    )
            pair_counts = [int(item.get("request_pair_count", -1)) for item in rank_calls]
            if not contiguous_partition_refines(
                pair_counts,
                contract_request_counts,
            ):
                raise RuntimeError(
                    "worker prefill calls are not a contiguous refinement of "
                    f"driver calls on rank {rank}: worker={pair_counts} "
                    f"driver={contract_request_counts}"
                )
            pair_counts_by_rank[rank] = pair_counts
            call_counts_by_rank[str(rank)] = len(rank_calls)
        if len({tuple(counts) for counts in pair_counts_by_rank.values()}) != 1:
            raise RuntimeError(f"TP prefill call partition mismatch: {pair_counts_by_rank}")
        return {
            "verified": True,
            "backend": "vllm",
            "mode": task.visual_token_shift,
            "model_key": model.key,
            "model_family": model_family,
            "matrix": self.matrix_name,
            "model_path": model.model_path,
            "implementation_sha256": inference_provenance["implementation_sha256"],
            "inference_fingerprint": inference_fingerprint,
            "run_id": run_id,
            "expected_tp_ranks": sorted(expected_ranks),
            "worker_handshake_count": len(handshakes),
            "shift_record_count": len(records),
            "runtime_contract_count": len(contracts),
            "validated_request_count": contract_request_count,
            "validated_prefill_calls_by_rank": call_counts_by_rank,
            "validated_prefill_pair_counts_by_rank": {
                str(rank): counts for rank, counts in pair_counts_by_rank.items()
            },
            "actual_vllm_engine_mode": contracts[0]["actual_vllm_engine_mode"],
            "actual_vllm_engine_module": contracts[0]["actual_vllm_engine_module"],
            "actual_vllm_engine_qualname": contracts[0]["actual_vllm_engine_qualname"],
            "actual_vllm_runtime_config": runtime_config,
        }

    def run_infer(self, task: Task, model: ModelSpec, env: dict[str, str], expected_count: int) -> int:
        log_path = self.log_root(task) / "infer" / f"{model.registry_name}_{task.dataset}_{self._timestamp()}.log"
        ensure_dir(self.model_output_root(task, model))
        ensure_dir(self.prediction_dir(task))
        batch_size = self.infer_batch_size_for_task(model, task)
        env["REPLAY_VISUAL_TOKEN_SHIFT_RUN_ID"] = (
            f"{self.matrix_name}.task{task.index}.pid{os.getpid()}.{time.time_ns()}"
        )
        inference_provenance = self.inference_provenance(task, model, env)
        env["REPLAY_INFERENCE_FINGERPRINT"] = inference_provenance["fingerprint"]
        smoke_certificate_evidence = None
        full_validation = truthy(
            env.get("REPLAY_VISUAL_TOKEN_SHIFT_FULL_VALIDATION", "0")
        )
        if task.visual_token_shift != "none" and not full_validation:
            try:
                smoke_certificate_evidence = (
                    self.verify_vllm_visual_token_shift_smoke_certificate(
                        task, model, env, inference_provenance
                    )
                )
            except Exception as exc:
                print(
                    f"[FAIL][INFER][SMOKE_CERTIFICATE] {task.tag}: "
                    f"{type(exc).__name__}: {exc}",
                    flush=True,
                )
                return 88
        cmd = [
            self.env_profiles[model.env_profile].python,
            "run.py",
            "--data",
            task.dataset,
            "--model",
            model.registry_name,
            "--work-dir",
            str(self.task_root(task)),
            "--mode",
            "infer",
            "--verbose",
            "--batch-size",
            str(batch_size),
            "--pred-output-dir",
            str(self.prediction_dir(task)),
            "--no-link-predictions",
        ]
        infer_nproc = str(env.get("VLMEVAL_INFER_NPROC", "")).strip()
        if infer_nproc:
            cmd.extend(["--api-nproc", infer_nproc])
        if self.resume_infer:
            cmd.append("--reuse")
        print(
            f"[START][INFER] {task.tag} model={model.registry_name} path={model.model_path} "
            f"gpus={env['CUDA_VISIBLE_DEVICES']} batch={batch_size}",
            flush=True,
        )
        rc = self.run_subprocess(cmd, env, log_path)
        post_run_provenance = self.inference_provenance(task, model, env)
        if post_run_provenance["fingerprint"] != inference_provenance["fingerprint"]:
            print(
                f"[FAIL][INFER][PROVENANCE] {task.tag}: inference inputs changed while the task was running",
                flush=True,
            )
            rc = 87
        if rc == 0 and smoke_certificate_evidence is not None:
            try:
                post_run_certificate_evidence = (
                    self.verify_vllm_visual_token_shift_smoke_certificate(
                        task, model, env, post_run_provenance
                    )
                )
                if post_run_certificate_evidence != smoke_certificate_evidence:
                    raise RuntimeError(
                        "smoke certificate identity changed while inference was running"
                    )
            except Exception as exc:
                print(
                    f"[FAIL][INFER][SMOKE_CERTIFICATE] {task.tag}: "
                    f"{type(exc).__name__}: {exc}",
                    flush=True,
                )
                rc = 88
        mechanism_evidence = None
        if rc == 0 and (
            task.visual_token_shift != "none"
            or truthy(env.get("REPLAY_VLLM_MATCHED_TOKEN_ROLL_RUNTIME", "0"))
        ):
            try:
                mechanism_evidence = self.verify_vllm_visual_token_shift_run(
                    task,
                    model,
                    env,
                    expected_count,
                    inference_provenance,
                )
                if smoke_certificate_evidence is not None:
                    mechanism_evidence["smoke_certificate"] = (
                        smoke_certificate_evidence
                    )
            except Exception as exc:
                print(
                    f"[FAIL][INFER][TOKEN_SHIFT] {task.tag}: {type(exc).__name__}: {exc}",
                    flush=True,
                )
                rc = 86
        if rc == 0:
            pred_file = self.infer_file_path(task, model)
            if self.prediction_file_valid(pred_file, expected_count):
                self.write_prediction_manifest(
                    task,
                    model,
                    expected_count,
                    pred_file,
                    mechanism_evidence=mechanism_evidence,
                    runtime_contract_run_id=env.get("REPLAY_VISUAL_TOKEN_SHIFT_RUN_ID"),
                    inference_provenance=inference_provenance,
                )
            print(f"[DONE][INFER] {task.tag} log={log_path}", flush=True)
        else:
            print(f"[FAIL][INFER] {task.tag} rc={rc} log={log_path}", flush=True)
        return rc

    def run_eval(self, task: Task, model: ModelSpec, env: dict[str, str], expected_count: int) -> int:
        log_path = self.log_root(task) / "eval" / f"{model.registry_name}_{task.dataset}_{self._timestamp()}.log"
        pred_file = self.infer_file_path(task, model)
        if pred_file is None:
            print(f"[FAIL][EVAL] {task.tag}: missing fixed prediction file", flush=True)
            return 1
        ensure_dir(self.eval_output_dir(task))
        cmd = [
            self.env_profiles[model.env_profile].python,
            "run.py",
            "--data",
            task.dataset,
            "--model",
            model.registry_name,
            "--work-dir",
            str(self.task_root(task)),
            "--mode",
            "eval",
            "--pred-file",
            str(pred_file),
            "--eval-dir",
            str(self.eval_output_dir(task)),
            "--no-link-predictions",
            "--api-nproc",
            str(self.evaluation_cfg.get("nproc", 8)),
            "--verbose",
            "--judge",
            str(self.evaluation_cfg.get("judge", "gpt-4o-mini")),
        ]
        print(f"[START][EVAL] {task.tag}", flush=True)
        rc = self.run_subprocess(cmd, env, log_path)
        if rc == 0:
            self.write_eval_manifest(task, model, expected_count, pred_file, rc)
            print(f"[DONE][EVAL] {task.tag} log={log_path}", flush=True)
        else:
            self.write_eval_manifest(task, model, expected_count, pred_file, rc)
            print(f"[FAIL][EVAL] {task.tag} rc={rc} log={log_path}", flush=True)
        return rc

    def run_single_task(self, model: ModelSpec, task: Task, gpu_ids: list[str]) -> None:
        env = self.build_env(model, task, gpu_ids)
        expected_count = self.get_expected_count(model, env, task.dataset)
        print(f"[TASK][START] {task.tag} expected={expected_count}", flush=True)
        if expected_count < 0:
            raise RuntimeError(f"dataset unavailable/build failed: {task.dataset}")

        infer_was_complete = (
            self.existing_infer_complete(task, model, expected_count, env)
            if self.inference_launch_mode == "existing"
            else self.infer_complete(task, model, expected_count, env)
        )
        if self.inference_launch_mode == "existing" and not infer_was_complete:
            raise RuntimeError(
                "existing inference failed producer provenance/certificate validation"
            )
        if infer_was_complete and self.eval_complete(task, model, expected_count):
            acc_paths = ", ".join(str(path) for path in self.acc_marker_paths(task, model))
            if acc_paths:
                print(f"[SKIP][TASK] {task.tag}: infer+eval complete via acc marker(s): {acc_paths}", flush=True)
            else:
                print(f"[SKIP][TASK] {task.tag}: infer+eval complete", flush=True)
            return

        if infer_was_complete:
            print(f"[SKIP][INFER] {task.tag}: complete", flush=True)
        else:
            if self.eval_artifacts(task, model):
                print(f"[CLEAN][EVAL] {task.tag}: remove stale eval artifacts before infer rerun", flush=True)
                self.cleanup_eval_artifacts(task, model)
            if self.infer_artifacts(task, model):
                if self.resume_infer:
                    if self.resume_infer_start_delay_seconds > 0:
                        print(
                            f"[WAIT][RESUME] {task.tag}: sleep {self.resume_infer_start_delay_seconds}s before reusing partial artifacts",
                            flush=True,
                        )
                        time.sleep(self.resume_infer_start_delay_seconds)
                    print(f"[RESUME][INFER] {task.tag}: reuse partial artifacts", flush=True)
                else:
                    print(f"[CLEAN][INFER+EVAL] {task.tag}: remove stale artifacts", flush=True)
                    self.cleanup_all_artifacts(task, model)
            infer_rc = self.run_infer(task, model, env, expected_count)
            if infer_rc != 0:
                raise RuntimeError(f"infer failed rc={infer_rc}")

        infer_still_complete = (
            self.existing_infer_complete(task, model, expected_count, env)
            if self.inference_launch_mode == "existing"
            else self.infer_complete(task, model, expected_count, env)
        )
        if not infer_still_complete:
            raise RuntimeError("infer incomplete after run")

        if self.eval_complete(task, model, expected_count):
            print(f"[SKIP][EVAL] {task.tag}: complete", flush=True)
            return
        if self.eval_artifacts(task, model):
            print(f"[CLEAN][EVAL] {task.tag}: remove stale eval artifacts", flush=True)
            self.cleanup_eval_artifacts(task, model)
        self.run_answer_format(task, model, env)

        launch_mode = str(self.evaluation_cfg.get("launch_mode", "fg")).lower()
        if launch_mode == "skip":
            print(f"[SKIP][EVAL] {task.tag}: launch_mode=skip", flush=True)
            return
        if launch_mode != "fg":
            raise ValueError(f"Unsupported evaluation launch mode: {launch_mode}")
        eval_rc = self.run_eval(task, model, env, expected_count)
        if eval_rc != 0:
            raise RuntimeError(f"eval failed rc={eval_rc}")
        if not self.eval_complete(task, model, expected_count):
            raise RuntimeError("eval incomplete after run")

    @staticmethod
    def _timestamp() -> str:
        from datetime import datetime

        return datetime.now().strftime("%Y%m%d%H%M%S")


def default_repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def parse_args() -> argparse.Namespace:
    repo_root = default_repo_root()
    parser = argparse.ArgumentParser(description="Unified replay benchmark launcher.")
    parser.add_argument("--matrix-config", type=Path, required=True)
    parser.add_argument("--model-config", type=Path, default=repo_root / "configs" / "models.yaml")
    parser.add_argument("--nodes", type=int, default=detect_num_nodes())
    parser.add_argument("--node-rank", type=int, default=detect_node_rank())
    parser.add_argument("--gpu-ids", type=str, default="")
    parser.add_argument("--models", type=str, default="")
    parser.add_argument("--policies", type=str, default="")
    parser.add_argument("--modes", type=str, default="")
    parser.add_argument("--transforms", type=str, default="")
    parser.add_argument("--visual-token-shifts", type=str, default="")
    parser.add_argument("--datasets", type=str, default="")
    parser.add_argument("--task-manifest", type=Path, default=None)
    parser.add_argument(
        "--scheduler",
        choices=("model_sequential", "gpu_pool"),
        default="",
        help="Task scheduler. model_sequential preserves legacy per-model workers; gpu_pool dynamically packs mixed TP sizes.",
    )
    parser.add_argument(
        "--manifest-is-node-shard",
        action="store_true",
        help="Treat --task-manifest as the already-sharded task list for --node-rank; do not split it again.",
    )
    parser.add_argument("--resume-infer", action="store_true", dest="resume_infer")
    parser.add_argument("--no-resume-infer", action="store_false", dest="resume_infer")
    parser.add_argument("--plan-only", action="store_true")
    parser.set_defaults(resume_infer=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    runner = BenchmarkRunner(default_repo_root(), args)
    return runner.run()


if __name__ == "__main__":
    raise SystemExit(main())
