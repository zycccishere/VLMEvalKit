#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
from typing import Any


# Allow one or more backslashes before "boxed" to tolerate escaped logs
# like "\\boxed{A}" while still enforcing answer-only format.
BOXED_ONLY_PATTERN = re.compile(r"^\\+boxed\{[^{}\n]+\}$")
ANSWER_BLOCK_PATTERN = re.compile(r"<\s*ANSWER\s*>(.*?)<\s*/\s*ANSWER\s*>", re.IGNORECASE | re.DOTALL)
ANSWER_OPEN_PATTERN = re.compile(r"<\s*ANSWER\s*>", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Check whether responses follow answer-only format and save a ratio report."
    )
    parser.add_argument("--pred-file", type=str, required=True, help="Prediction file (.xlsx/.tsv/.csv).")
    parser.add_argument("--out-json", type=str, required=True, help="Output summary json path.")
    parser.add_argument("--out-fail-jsonl", type=str, default="", help="Optional jsonl path of failed samples.")
    parser.add_argument("--response-col", type=str, default="prediction", help="Primary response column name.")
    parser.add_argument("--fallback-col", type=str, default="detailed_prediction", help="Fallback response column name.")
    parser.add_argument(
        "--require-boxed",
        type=int,
        default=0,
        help="If 1, require strict one-line '\\\\boxed{...}' only format. If 0, only require single-line output.",
    )
    parser.add_argument("--max-fails", type=int, default=50, help="Max failed examples stored in summary.")
    return parser.parse_args()


def load_rows(path: str) -> list[dict[str, Any]]:
    lower = path.lower()
    if lower.endswith(".tsv") or lower.endswith(".csv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with open(path, "r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f, delimiter=delimiter))

    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(x) if x is not None else "" for x in rows[0]]
        data_rows = []
        for row in rows[1:]:
            item = {}
            for i, col in enumerate(header):
                if col:
                    item[col] = row[i] if i < len(row) else None
            data_rows.append(item)
        return data_rows

    raise ValueError(f"Unsupported file extension: {path}")


def pick_response(row: dict[str, Any], response_col: str, fallback_col: str) -> str:
    value = row.get(response_col, None)
    if value is None or str(value).strip() == "":
        value = row.get(fallback_col, "")
    return str(value if value is not None else "").strip()


def extract_answer_text(response: str) -> str:
    text = str(response).strip()
    if text == "":
        return ""

    # Convert escaped newlines in serialized logs.
    text = text.replace("\\n", "\n")

    # Prefer explicit <ANSWER>...</ANSWER> region if present.
    m = ANSWER_BLOCK_PATTERN.search(text)
    if m is not None:
        text = m.group(1).strip()
    else:
        # Tolerate missing closing tag: keep content after opening <ANSWER>.
        m_open = ANSWER_OPEN_PATTERN.search(text)
        if m_open is not None:
            text = text[m_open.end() :].strip()
        text = re.sub(r"<\s*/\s*ANSWER\s*>", "", text, flags=re.IGNORECASE).strip()

    return text


def normalize_response(response: str) -> str:
    text = extract_answer_text(response)
    if text == "":
        return ""

    # Keep the first non-empty logical line. This accepts cases like:
    # <ANSWER>\n27\n</ANSWER>
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return ""
    return lines[0]


def check_format(response: str, require_boxed: bool) -> tuple[bool, str, str]:
    extracted = extract_answer_text(response)
    normalized = normalize_response(response)
    if normalized == "":
        return False, "empty"
    logical_lines = [ln.strip() for ln in extracted.splitlines() if ln.strip()]
    if not logical_lines:
        return False, "empty", normalized
    if require_boxed:
        if BOXED_ONLY_PATTERN.fullmatch(normalized):
            return True, "ok", normalized
        return False, "not_boxed_only", normalized
    if len(logical_lines) != 1:
        return False, "not_single_line", normalized
    return True, "ok", normalized


def main() -> None:
    args = parse_args()
    rows = load_rows(args.pred_file)
    require_boxed = bool(args.require_boxed)

    total = len(rows)
    ok_count = 0
    reason_counts: dict[str, int] = {}
    fail_examples = []
    fail_jsonl_records = []

    for ridx, row in enumerate(rows):
        response = pick_response(row, args.response_col, args.fallback_col)
        is_ok, reason, normalized = check_format(response, require_boxed=require_boxed)
        if is_ok:
            ok_count += 1
            continue

        reason_counts[reason] = reason_counts.get(reason, 0) + 1
        sample_idx = row.get("index", ridx)
        fail_rec = {
            "row": ridx,
            "index": sample_idx,
            "reason": reason,
            "response": response,
            "normalized": normalized,
        }
        fail_jsonl_records.append(fail_rec)
        if len(fail_examples) < max(args.max_fails, 0):
            fail_examples.append(fail_rec)

    ratio = (ok_count / total) if total > 0 else 0.0
    report = {
        "pred_file": os.path.abspath(args.pred_file),
        "total_samples": total,
        "correct_format_samples": ok_count,
        "correct_format_ratio": ratio,
        "response_column": args.response_col,
        "fallback_column": args.fallback_col,
        "require_boxed": require_boxed,
        "format_mode": "boxed_only" if require_boxed else "single_line",
        "pattern": BOXED_ONLY_PATTERN.pattern if require_boxed else "single_line",
        "fail_reason_counts": reason_counts,
        "fail_examples": fail_examples,
    }

    out_dir = os.path.dirname(os.path.abspath(args.out_json))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    if args.out_fail_jsonl.strip():
        fail_dir = os.path.dirname(os.path.abspath(args.out_fail_jsonl))
        if fail_dir:
            os.makedirs(fail_dir, exist_ok=True)
        with open(args.out_fail_jsonl, "w", encoding="utf-8") as f:
            for rec in fail_jsonl_records:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    print(
        (
            f"[answer-format] total={total} correct={ok_count} ratio={ratio:.4f} "
            f"mode={'boxed_only' if require_boxed else 'single_line'} "
            f"require_boxed={int(require_boxed)} out={args.out_json}"
        ),
        flush=True,
    )


if __name__ == "__main__":
    main()
