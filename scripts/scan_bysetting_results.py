#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any


DATASET_COLUMNS = [
    "DynaMath",
    "MathVision",
    "MathVista_MINI",
    "LogicVista",
    "VisualPuzzles",
    "VisuLogic",
    "AI2D_TEST",
    "OCRBench",
    "SEEDBench2_Plus",
]

MODEL_KEY_TO_LABEL = {
    "qwen35_35b_a3b": "Qwen3.5 35B-A3B",
    "qwen35_27b": "Qwen3.5 27B",
    "qwen35_9b": "Qwen3.5 9B",
    "qwen35_4b": "Qwen3.5 4B",
    "qwen2vl_7b": "Qwen2 7B",
    "qwen25vl_7b": "Qwen2.5 7B",
    "minicpm_v_45": "MiniCPM-V",
    "minicpm_o_45": "MiniCPM-o",
}

KNOWN_DATASETS = sorted(DATASET_COLUMNS, key=len, reverse=True)


@dataclass
class MetricRecord:
    policy: str
    setting: str
    model_key: str
    model_label: str
    registry_name: str
    dataset: str
    value: float
    kind: str
    priority: int
    source_path: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scan runs/by_setting and summarize metrics plus infer-complete zero fills.")
    parser.add_argument("--runs-root", type=Path, required=True)
    parser.add_argument("--out-json", type=Path, required=True)
    return parser.parse_args()


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().strip('"')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def maybe_pct(value: float) -> float:
    return value * 100.0 if abs(value) <= 1.0 else value


def detect_dataset(path: Path) -> str | None:
    text = path.name
    for dataset in KNOWN_DATASETS:
        if dataset in text:
            return dataset
    return None


def choose_overall_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    preferred = {"overall", "average", "none"}
    for row in rows:
        for value in row.values():
            if str(value).strip().lower() in preferred:
                return row
    return rows[0]


def parse_csv_metric(path: Path, dataset_hint: str | None) -> tuple[str, float, int] | None:
    with path.open("r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        return None
    lower_headers = {key.lower(): key for key in rows[0].keys()}

    if "dataset" in lower_headers and "overall" in lower_headers:
        dataset = rows[0][lower_headers["dataset"]]
        value = to_float(rows[0][lower_headers["overall"]])
        if dataset and value is not None:
            return dataset, value, 10
        return None

    dataset = dataset_hint or detect_dataset(path)
    if not dataset:
        return None

    if "setting" in lower_headers and "overall" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["overall"]])
        if value is not None:
            return dataset, maybe_pct(value), 30

    if "overall" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["overall"]])
        if value is not None:
            return dataset, maybe_pct(value), 20

    if "acc" in lower_headers:
        row = choose_overall_row(rows)
        value = to_float(row[lower_headers["acc"]])
        if value is not None:
            return dataset, maybe_pct(value), 20

    return None


def parse_json_metric(path: Path, dataset_hint: str | None) -> tuple[str, float, int] | None:
    dataset = dataset_hint or detect_dataset(path)
    if not dataset:
        return None
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if dataset == "OCRBench":
        value = to_float(data.get("Final Score Norm"))
        if value is not None:
            return dataset, value, 30
    overall = to_float(data.get("Overall"))
    if overall is not None:
        return dataset, maybe_pct(overall), 30
    return None


def parse_metric(path: Path) -> tuple[str, float, int] | None:
    dataset_hint = detect_dataset(path)
    if path.suffix == ".json":
        return parse_json_metric(path, dataset_hint)
    if path.suffix == ".csv":
        return parse_csv_metric(path, dataset_hint)
    return None


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    lower = path.name.lower()
    if lower.endswith(".tsv") or lower.endswith(".csv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh, delimiter=delimiter))
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(v) if v is not None else "" for v in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for idx, col in enumerate(header):
                if col:
                    item[col] = row[idx] if idx < len(row) else None
            out.append(item)
        return out
    raise ValueError(f"Unsupported tabular file: {path}")


def infer_complete(pred_file: Path | None, expected: int) -> bool:
    if pred_file is None or expected < 0 or not pred_file.exists():
        return False
    try:
        rows = load_tabular_rows(pred_file)
    except Exception:
        return False
    if len(rows) != expected:
        return False
    candidate_cols = [col for col in ("prediction", "description", "detailed_prediction") if rows and col in rows[0]]
    if not candidate_cols:
        return True
    saw_nonblank = False
    for row in rows:
        values = [row.get(col) for col in candidate_cols]
        if not all(is_blank(value) for value in values):
            saw_nonblank = True
        desc = str(row.get("description", "")).strip()
        if desc.startswith("[FAILED_INFER]") or "Failed to obtain answer via API." in desc:
            return False
    return saw_nonblank


def get_expected_count(name: str) -> int:
    import contextlib
    import io

    from vlmeval.dataset import build_dataset

    buf = io.StringIO()
    dataset = None
    err = None
    with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
        try:
            dataset = build_dataset(name)
        except Exception as exc:
            err = exc
    if dataset is None or err is not None:
        return -1
    try:
        return int(len(dataset))
    except Exception:
        data = getattr(dataset, "data", None)
        return int(len(data)) if data is not None else -1


def collect_metric_records(runs_root: Path) -> dict[tuple[str, str, str, str], MetricRecord]:
    best: dict[tuple[str, str, str, str], MetricRecord] = {}
    for path in sorted(runs_root.rglob("*")):
        if not path.is_file():
            continue
        if not path.name.endswith(("_acc.csv", "_score.csv", "_score.json")):
            continue
        rel = path.relative_to(runs_root)
        if len(rel.parts) < 5:
            continue
        policy, setting, model_key, registry_name = rel.parts[0], rel.parts[1], rel.parts[2], rel.parts[3]
        if policy not in {"direct", "default"}:
            continue
        model_label = MODEL_KEY_TO_LABEL.get(model_key)
        if not model_label:
            continue
        parsed = parse_metric(path)
        if not parsed:
            continue
        dataset, value, priority = parsed
        if dataset not in DATASET_COLUMNS:
            continue
        key = (policy, model_label, setting, dataset)
        record = MetricRecord(
            policy=policy,
            setting=setting,
            model_key=model_key,
            model_label=model_label,
            registry_name=registry_name,
            dataset=dataset,
            value=value,
            kind="metric",
            priority=priority,
            source_path=str(path),
        )
        prev = best.get(key)
        if prev is None or (record.priority, record.source_path) > (prev.priority, prev.source_path):
            best[key] = record
    return best


def collect_infer_zero_records(
    runs_root: Path,
    metric_records: dict[tuple[str, str, str, str], MetricRecord],
) -> list[MetricRecord]:
    expected_counts = {dataset: get_expected_count(dataset) for dataset in DATASET_COLUMNS}
    out: list[MetricRecord] = []
    for policy_dir in sorted(path for path in runs_root.iterdir() if path.is_dir()):
        policy = policy_dir.name
        if policy not in {"direct", "default"}:
            continue
        for setting_dir in sorted(path for path in policy_dir.iterdir() if path.is_dir()):
            setting = setting_dir.name
            for model_dir in sorted(path for path in setting_dir.iterdir() if path.is_dir()):
                model_key = model_dir.name
                model_label = MODEL_KEY_TO_LABEL.get(model_key)
                if not model_label:
                    continue
                for registry_dir in sorted(path for path in model_dir.iterdir() if path.is_dir() and not path.name.startswith("_")):
                    registry_name = registry_dir.name
                    for dataset in DATASET_COLUMNS:
                        key = (policy, model_label, setting, dataset)
                        if key in metric_records:
                            continue
                        expected = expected_counts.get(dataset, -1)
                        xlsx = registry_dir / f"{registry_name}_{dataset}.xlsx"
                        tsv = registry_dir / f"{registry_name}_{dataset}.tsv"
                        pred_file = xlsx if xlsx.exists() else tsv if tsv.exists() else None
                        if not infer_complete(pred_file, expected):
                            continue
                        out.append(
                            MetricRecord(
                                policy=policy,
                                setting=setting,
                                model_key=model_key,
                                model_label=model_label,
                                registry_name=registry_name,
                                dataset=dataset,
                                value=0.0,
                                kind="infer_zero",
                                priority=0,
                                source_path=str(pred_file),
                            )
                        )
    return out


def main() -> int:
    args = parse_args()
    metric_records = collect_metric_records(args.runs_root)
    infer_zero_records = collect_infer_zero_records(args.runs_root, metric_records)
    combined = list(metric_records.values()) + infer_zero_records
    combined.sort(key=lambda r: (r.policy, r.model_label, r.setting, r.dataset, r.kind, r.source_path))
    payload = {
        "runs_root": str(args.runs_root),
        "record_count": len(combined),
        "metric_count": len(metric_records),
        "infer_zero_count": len(infer_zero_records),
        "records": [
            {
                "policy": r.policy,
                "setting": r.setting,
                "model_key": r.model_key,
                "model_label": r.model_label,
                "registry_name": r.registry_name,
                "dataset": r.dataset,
                "value": r.value,
                "kind": r.kind,
                "source_path": r.source_path,
            }
            for r in combined
        ],
    }
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: payload[k] for k in ("runs_root", "record_count", "metric_count", "infer_zero_count")}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
