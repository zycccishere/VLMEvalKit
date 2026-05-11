#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import shutil
from dataclasses import dataclass
from pathlib import Path

import yaml


EXPECTED_ROWS = {"DynaMath": 5010}


@dataclass(frozen=True)
class ModelSpec:
    key: str
    display_name: str
    registry_name: str


def load_yaml(path: Path) -> dict:
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"Expected mapping YAML: {path}")
    return data


def split_names(raw: str) -> list[str]:
    if not raw:
        return []
    return [part for part in raw.replace(",", " ").split() if part]


def count_rows(path: Path) -> int:
    lower = path.name.lower()
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return max(ws.max_row - 1, 0)
    if lower.endswith(".tsv"):
        import csv

        with path.open("r", encoding="utf-8", newline="") as fh:
            rows = list(csv.reader(fh, delimiter="\t"))
        return max(len(rows) - 1, 0)
    return -1


def is_valid_prediction_file(path: Path, dataset: str) -> bool:
    if not path.is_file() or path.stat().st_size <= 0:
        return False
    expected = EXPECTED_ROWS.get(dataset)
    if expected is None:
        return True
    try:
        return count_rows(path) == expected
    except Exception:
        return False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed infer outputs from legacy runs.")
    parser.add_argument("--matrix-config", required=True)
    parser.add_argument("--model-config", required=True)
    parser.add_argument("--source-root", action="append", default=[])
    parser.add_argument("--link-mode", choices=["auto", "hardlink", "copy"], default="auto")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def resolve_results_root(repo_root: Path, raw: str) -> Path:
    root = Path(raw)
    return root if root.is_absolute() else repo_root / root


def default_source_roots(repo_root: Path) -> list[Path]:
    return [
        repo_root / "runs" / "standard",
        repo_root / "runs" / "by_setting_core4_current_models",
        repo_root / "runs" / "by_setting",
    ]


def parse_standard_candidate(path: Path, display_to_key: dict[str, str]) -> tuple[str, str, str] | None:
    lower = str(path).lower()
    if "direct" in lower:
        return None
    task_dir = None
    for part in path.parts:
        if "__" in part and part.endswith("__last1"):
            task_dir = part
            break
    if task_dir is None:
        return None
    pieces = task_dir.split("__", 2)
    if len(pieces) != 3:
        return None
    display_name, mode, last_flag = pieces
    if last_flag != "last1":
        return None
    model_key = display_to_key.get(display_name)
    if model_key is None:
        return None
    dataset = path.stem.rsplit("_", 1)[-1]
    return model_key, mode, dataset


def parse_by_setting_candidate(path: Path, source_root: Path, model_keys: set[str]) -> tuple[str, str, str] | None:
    rel = path.relative_to(source_root)
    if len(rel.parts) < 5:
        return None
    policy, mode, model_key = rel.parts[0], rel.parts[1], rel.parts[2]
    if policy != "default" or model_key not in model_keys:
        return None
    dataset = path.stem.rsplit("_", 1)[-1]
    return model_key, mode, dataset


def build_source_index(
    source_roots: list[Path],
    display_to_key: dict[str, str],
    model_keys: set[str],
) -> dict[tuple[str, str, str], list[Path]]:
    index: dict[tuple[str, str, str], list[Path]] = {}
    for source_root in source_roots:
        if not source_root.exists():
            continue
        for path in source_root.rglob("*_DynaMath.xlsx"):
            parsed = None
            if source_root.name == "standard":
                parsed = parse_standard_candidate(path, display_to_key)
            else:
                parsed = parse_by_setting_candidate(path, source_root, model_keys)
            if parsed is None:
                continue
            if not is_valid_prediction_file(path, parsed[2]):
                continue
            index.setdefault(parsed, []).append(path)
    for key in index:
        index[key].sort(key=lambda p: (p.stat().st_mtime, str(p)), reverse=True)
    return index


def link_or_copy(src: Path, dst: Path, mode: str) -> str:
    dst.parent.mkdir(parents=True, exist_ok=True)
    if mode in {"auto", "hardlink"}:
        try:
            os.link(src, dst)
            return "hardlink"
        except OSError:
            if mode == "hardlink":
                raise
    shutil.copy2(src, dst)
    return "copy"


def main() -> int:
    args = parse_args()
    matrix = load_yaml(Path(args.matrix_config))
    models_cfg = load_yaml(Path(args.model_config))
    repo_root = Path(matrix["repo_root"])
    results_root = resolve_results_root(repo_root, str(matrix["results_root"]))
    model_specs = {
        key: ModelSpec(
            key=key,
            display_name=str(raw["display_name"]),
            registry_name=str(raw["registry_name"]),
        )
        for key, raw in models_cfg["models"].items()
    }
    source_roots = [Path(p) for p in args.source_root] if args.source_root else default_source_roots(repo_root)
    display_to_key = {spec.display_name: spec.key for spec in model_specs.values()}
    source_index = build_source_index(source_roots, display_to_key, set(model_specs))

    tasks = []
    for model_key in matrix["models"]:
        spec = model_specs[model_key]
        for policy in matrix["policies"]:
            if policy != "default":
                continue
            for mode in matrix["replay_modes"]:
                for dataset in matrix["datasets"]:
                    target = results_root / policy / mode / model_key / spec.registry_name / f"{spec.registry_name}_{dataset}.xlsx"
                    tasks.append((model_key, mode, dataset, target))

    seeded = []
    skipped_existing = []
    missing = []
    for model_key, mode, dataset, target in tasks:
        if is_valid_prediction_file(target, dataset):
            skipped_existing.append(str(target))
            continue
        candidates = source_index.get((model_key, mode, dataset), [])
        if not candidates:
            missing.append({"model_key": model_key, "mode": mode, "dataset": dataset})
            continue
        src = candidates[0]
        if args.dry_run:
            seeded.append(
                {
                    "model_key": model_key,
                    "mode": mode,
                    "dataset": dataset,
                    "src": str(src),
                    "dst": str(target),
                    "action": "dry-run",
                }
            )
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists():
            target.unlink()
        method = link_or_copy(src, target, args.link_mode)
        seeded.append(
            {
                "model_key": model_key,
                "mode": mode,
                "dataset": dataset,
                "src": str(src),
                "dst": str(target),
                "action": method,
            }
        )

    print(
        json.dumps(
            {
                "results_root": str(results_root),
                "source_roots": [str(p) for p in source_roots],
                "seeded_count": len(seeded),
                "skipped_existing_count": len(skipped_existing),
                "missing_count": len(missing),
                "seeded": seeded,
                "missing": missing,
            },
            ensure_ascii=False,
            indent=2,
        ),
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
