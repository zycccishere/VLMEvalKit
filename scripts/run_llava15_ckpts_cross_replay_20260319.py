#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import pickle
import queue
import shutil
import subprocess
import threading
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


def truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def split_names(raw: str) -> list[str]:
    if not raw:
        return []
    return [part for part in raw.replace(",", " ").split() if part]


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    lower = path.name.lower()
    if lower.endswith(".tsv") or lower.endswith(".csv"):
        delimiter = "\t" if lower.endswith(".tsv") else ","
        with path.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh, delimiter=delimiter))
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(v) if v is not None else "" for v in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for idx, col in enumerate(header):
                if col:
                    item[col] = row[idx] if idx < len(row) else None
            out.append(item)
        return out
    raise ValueError(f"Unsupported tabular file: {path}")


@dataclass(frozen=True)
class Task:
    index: int
    train_mode: str
    eval_mode: str
    dataset: str

    @property
    def tag(self) -> str:
        return f"{self.train_mode}_train__{self.eval_mode}_eval__{self.dataset}"


class LlavaCrossReplayRunner:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.repo_root = Path(args.repo_root).resolve()
        self.python_bin = Path(args.python_bin)
        self.work_root = Path(args.work_root)
        self.ckpt_root = Path(args.ckpt_root)
        self.model_name = args.model_name
        self.judge = args.judge
        self.nproc = args.nproc
        self.eval_workers = args.eval_workers
        self.batch_size = args.batch_size
        self.resume_infer = args.resume_infer
        self.run_mode = args.mode
        self.gpu_ids = split_names(args.gpu_ids)
        if not self.gpu_ids:
            raise SystemExit("No GPU ids provided.")
        self.datasets = split_names(args.datasets)
        if not self.datasets:
            raise SystemExit("No datasets provided.")
        self.combos = self._build_combos()
        self.tasks = self._build_tasks()
        self.expected_count_cache: dict[str, int] = {}

    def _build_combos(self) -> list[tuple[str, str]]:
        return [
            ("image_text", "image_text"),
            ("image_text", "text_image"),
            ("image_text", "image_text_text"),
            ("image_text_text", "image_text_text"),
            ("image_text", "image_text_image"),
            ("image_text_image", "image_text_image"),
            ("image_text", "image_text_image_text"),
            ("image_text_image_text", "image_text_image_text"),
            ("image_text", "image_image_text"),
        ]

    def _build_tasks(self) -> list[Task]:
        tasks: list[Task] = []
        idx = 0
        for train_mode, eval_mode in self.combos:
            for dataset in self.datasets:
                tasks.append(Task(index=idx, train_mode=train_mode, eval_mode=eval_mode, dataset=dataset))
                idx += 1
        return tasks

    def ckpt_path(self, train_mode: str) -> Path:
        path = self.ckpt_root / f"llava-v1.5-13b-{train_mode}"
        if not path.is_dir():
            raise FileNotFoundError(f"Missing checkpoint for train mode '{train_mode}': {path}")
        return path

    def task_root(self, task: Task) -> Path:
        ckpt_name = f"llava-v1.5-13b-{task.train_mode}"
        tag = f"{task.train_mode}_train__{task.eval_mode}_eval"
        return self.work_root / ckpt_name / tag

    def model_output_root(self, task: Task) -> Path:
        return self.task_root(task) / self.model_name

    def log_root(self, task: Task) -> Path:
        return self.task_root(task) / "_logs"

    def infer_file_path(self, task: Task) -> Path | None:
        model_dir = self.model_output_root(task)
        xlsx = model_dir / f"{self.model_name}_{task.dataset}.xlsx"
        tsv = model_dir / f"{self.model_name}_{task.dataset}.tsv"
        if xlsx.exists():
            return xlsx
        if tsv.exists():
            return tsv
        return None

    def infer_artifacts(self, task: Task) -> list[Path]:
        model_dir = self.model_output_root(task)
        if not model_dir.exists():
            return []
        return sorted(model_dir.glob(f"{self.model_name}_{task.dataset}*"))

    def infer_primary_paths(self, task: Task) -> set[Path]:
        model_dir = self.model_output_root(task)
        return {
            model_dir / f"{self.model_name}_{task.dataset}.xlsx",
            model_dir / f"{self.model_name}_{task.dataset}.tsv",
        }

    def acc_marker_paths(self, task: Task) -> list[Path]:
        model_dir = self.model_output_root(task)
        if not model_dir.exists():
            return []
        return sorted(model_dir.glob(f"*_{task.dataset}*_acc.csv"))

    def acc_complete(self, task: Task) -> bool:
        for path in self.acc_marker_paths(task):
            try:
                if path.is_file() and path.stat().st_size > 0:
                    return True
            except OSError:
                continue
        return False

    def eval_artifacts(self, task: Task) -> list[Path]:
        out: list[Path] = []
        primary = self.infer_primary_paths(task)
        for path in self.infer_artifacts(task):
            if path in primary:
                continue
            if path.name.endswith("_acc.csv"):
                continue
            if path.name.endswith("_answer_format_report.json") or path.name.endswith("_answer_format_failures.jsonl"):
                continue
            out.append(path)
        return out

    def cleanup_all_artifacts(self, task: Task) -> None:
        for path in self.infer_artifacts(task):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)

    def cleanup_eval_artifacts(self, task: Task) -> None:
        for path in self.eval_artifacts(task):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)

    def build_env(self, task: Task, gpu_id: str) -> dict[str, str]:
        env = dict(os.environ)
        env["CUDA_VISIBLE_DEVICES"] = gpu_id
        env["MODEL_PATH"] = str(self.ckpt_path(task.train_mode))
        env["REPLAY_MODE"] = task.eval_mode
        return env

    def get_expected_count(self, task: Task, env: dict[str, str]) -> int:
        if task.dataset in self.expected_count_cache:
            return self.expected_count_cache[task.dataset]
        code = """
import contextlib
import io
import sys
from vlmeval.dataset import build_dataset

name = sys.argv[1]
buf = io.StringIO()
dataset = None
err = None
with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
    try:
        dataset = build_dataset(name)
    except Exception as exc:
        err = exc
if dataset is None or err is not None:
    captured = buf.getvalue().strip()
    if captured:
        print(captured, file=sys.stderr)
    if err is not None:
        print(f"[get_expected_count] build_dataset({name}) failed: {err}", file=sys.stderr)
    print(-1)
    raise SystemExit(0)
print(int(len(dataset)))
"""
        proc = subprocess.run(
            [str(self.python_bin), "-c", code, task.dataset],
            cwd=self.repo_root,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        count = -1
        for line in proc.stdout.splitlines():
            line = line.strip()
            if line.lstrip("-").isdigit():
                count = int(line)
        self.expected_count_cache[task.dataset] = count
        return count

    def infer_complete(self, task: Task, expected: int) -> bool:
        if expected < 0:
            return False
        pred_file = self.infer_file_path(task)
        if pred_file is None:
            return False
        try:
            rows = load_tabular_rows(pred_file)
        except Exception:
            return False
        if len(rows) != expected:
            return False
        candidate_cols = [col for col in ("prediction", "description", "detailed_prediction") if rows and col in rows[0]]
        if not candidate_cols:
            return True
        for row in rows:
            values = [row.get(col) for col in candidate_cols]
            if all(is_blank(value) for value in values):
                return False
            desc = str(row.get("description", "")).strip()
            if desc.startswith("[FAILED_INFER]") or "Failed to obtain answer via API." in desc:
                return False
        return True

    def eval_complete(self, task: Task, expected: int) -> bool:
        if self.acc_complete(task):
            return True
        eval_files = self.eval_artifacts(task)
        if not eval_files:
            return False
        has_sample_pkl = False
        has_summary = False
        for path in eval_files:
            if path.name.endswith("_result.pkl"):
                has_sample_pkl = True
                try:
                    with path.open("rb") as fh:
                        obj = pickle.load(fh)
                    if not hasattr(obj, "__len__") or len(obj) != expected:
                        return False
                except Exception:
                    return False
            elif path.suffix == ".csv":
                try:
                    content = path.read_text(encoding="utf-8", errors="ignore").lower()
                except Exception:
                    return False
                if "overall" not in content:
                    return False
                has_summary = True
            elif path.suffix == ".json":
                if path.stat().st_size <= 0:
                    return False
                has_summary = True
        return has_sample_pkl or has_summary

    def run_subprocess(self, cmd: list[str], env: dict[str, str], log_path: Path) -> int:
        ensure_dir(log_path.parent)
        with log_path.open("w", encoding="utf-8") as fh:
            proc = subprocess.run(
                cmd,
                cwd=Path("/tmp"),
                env=env,
                stdout=fh,
                stderr=subprocess.STDOUT,
                text=True,
            )
        return proc.returncode

    def run_answer_format(self, task: Task, env: dict[str, str]) -> None:
        pred_file = self.infer_file_path(task)
        if pred_file is None:
            return
        model_dir = self.model_output_root(task)
        report = model_dir / f"{self.model_name}_{task.dataset}_answer_format_report.json"
        failures = model_dir / f"{self.model_name}_{task.dataset}_answer_format_failures.jsonl"
        log_path = self.log_root(task) / "answer_format" / f"{self.model_name}_{task.dataset}_{self._timestamp()}.log"
        cmd = [
            str(self.python_bin),
            str(self.repo_root / "scripts" / "postprocess_answer_format.py"),
            "--pred-file",
            str(pred_file),
            "--out-json",
            str(report),
            "--out-fail-jsonl",
            str(failures),
            "--response-col",
            "prediction",
            "--fallback-col",
            "detailed_prediction",
            "--require-boxed",
            "0",
            "--max-fails",
            "50",
        ]
        rc = self.run_subprocess(cmd, env, log_path)
        if rc != 0:
            print(f"[FAIL][FORMAT] {task.tag}: rc={rc} log={log_path}", flush=True)

    def run_infer(self, task: Task, env: dict[str, str]) -> int:
        log_path = self.log_root(task) / "infer" / f"{self.model_name}_{task.dataset}_{self._timestamp()}.log"
        ensure_dir(self.model_output_root(task))
        cmd = [
            str(self.python_bin),
            str(self.repo_root / "run.py"),
            "--data",
            task.dataset,
            "--model",
            self.model_name,
            "--work-dir",
            str(self.task_root(task)),
            "--mode",
            "infer",
            "--batch-size",
            str(self.batch_size),
            "--verbose",
        ]
        print(f"[START][INFER] {task.tag} gpu={env['CUDA_VISIBLE_DEVICES']} batch={self.batch_size}", flush=True)
        rc = self.run_subprocess(cmd, env, log_path)
        if rc == 0:
            print(f"[DONE][INFER] {task.tag} log={log_path}", flush=True)
        else:
            print(f"[FAIL][INFER] {task.tag} rc={rc} log={log_path}", flush=True)
        return rc

    def run_eval(self, task: Task, env: dict[str, str]) -> int:
        log_path = self.log_root(task) / "eval" / f"{self.model_name}_{task.dataset}_{self._timestamp()}.log"
        cmd = [
            str(self.python_bin),
            str(self.repo_root / "run.py"),
            "--data",
            task.dataset,
            "--model",
            self.model_name,
            "--work-dir",
            str(self.task_root(task)),
            "--mode",
            "eval",
            "--nproc",
            str(self.nproc),
            "--judge",
            self.judge,
            "--verbose",
        ]
        print(f"[START][EVAL] {task.tag}", flush=True)
        rc = self.run_subprocess(cmd, env, log_path)
        if rc == 0:
            print(f"[DONE][EVAL] {task.tag} log={log_path}", flush=True)
        else:
            print(f"[FAIL][EVAL] {task.tag} rc={rc} log={log_path}", flush=True)
        return rc

    def enqueue_eval_if_needed(self, task: Task, env: dict[str, str], expected: int, eval_queue: queue.Queue[tuple[Task, dict[str, str], int] | None]) -> None:
        if self.run_mode not in {"all", "eval"}:
            return
        eval_queue.put((task, dict(env), expected))

    def run_infer_task(self, task: Task, gpu_id: str, eval_queue: queue.Queue[tuple[Task, dict[str, str], int] | None]) -> None:
        if self.acc_complete(task):
            acc_paths = ", ".join(str(path) for path in self.acc_marker_paths(task))
            print(f"[SKIP][ACC] {task.tag}: complete via acc marker(s): {acc_paths}", flush=True)
            return

        env = self.build_env(task, gpu_id)
        expected = self.get_expected_count(task, env)
        print(f"[TASK][START] {task.tag} expected={expected}", flush=True)
        if expected < 0:
            print(f"[SKIP][DATASET] {task.tag}: unavailable/build failed", flush=True)
            return

        if self.run_mode == "eval":
            if not self.infer_complete(task, expected):
                print(f"[SKIP][EVAL] {task.tag}: infer incomplete", flush=True)
                return
            self.enqueue_eval_if_needed(task, env, expected, eval_queue)
            return

        if self.infer_complete(task, expected):
            print(f"[SKIP][INFER] {task.tag}: complete", flush=True)
        else:
            if self.infer_artifacts(task):
                if self.resume_infer:
                    print(f"[RESUME][INFER] {task.tag}: reuse partial artifacts", flush=True)
                else:
                    print(f"[CLEAN][INFER+EVAL] {task.tag}: remove stale artifacts", flush=True)
                    self.cleanup_all_artifacts(task)
            if self.run_infer(task, env) != 0:
                return

        if not self.infer_complete(task, expected):
            print(f"[SKIP][EVAL] {task.tag}: infer incomplete", flush=True)
            return

        self.enqueue_eval_if_needed(task, env, expected, eval_queue)

    def run_eval_task(self, task: Task, env: dict[str, str], expected: int) -> None:
        if self.acc_complete(task):
            acc_paths = ", ".join(str(path) for path in self.acc_marker_paths(task))
            print(f"[SKIP][ACC] {task.tag}: complete via acc marker(s): {acc_paths}", flush=True)
            return
        if not self.infer_complete(task, expected):
            print(f"[SKIP][EVAL] {task.tag}: infer incomplete", flush=True)
            return
        self.run_answer_format(task, env)
        if self.eval_complete(task, expected):
            print(f"[SKIP][EVAL] {task.tag}: complete", flush=True)
            return
        if self.eval_artifacts(task):
            print(f"[CLEAN][EVAL] {task.tag}: remove stale eval artifacts", flush=True)
            self.cleanup_eval_artifacts(task)
        self.run_eval(task, env)

    def run_infer_worker(
        self,
        worker_idx: int,
        gpu_id: str,
        task_queue: queue.Queue[Task],
        eval_queue: queue.Queue[tuple[Task, dict[str, str], int] | None],
    ) -> None:
        initial_qsize = task_queue.qsize()
        claimed = 0
        print(f"[WORKER][START][INFER] slot={worker_idx} gpu={gpu_id} queue={initial_qsize}", flush=True)
        while True:
            try:
                task = task_queue.get_nowait()
            except queue.Empty:
                break
            claimed += 1
            try:
                self.run_infer_task(task, gpu_id, eval_queue)
            except Exception as exc:
                print(f"[WORKER][FAIL] {task.tag}: {exc}", flush=True)
            finally:
                task_queue.task_done()
        print(f"[WORKER][DONE][INFER] slot={worker_idx} gpu={gpu_id} completed={claimed}", flush=True)

    def run_eval_worker(
        self,
        worker_idx: int,
        eval_queue: queue.Queue[tuple[Task, dict[str, str], int] | None],
    ) -> None:
        print(f"[WORKER][START][EVAL] slot={worker_idx}", flush=True)
        claimed = 0
        while True:
            item = eval_queue.get()
            if item is None:
                eval_queue.task_done()
                break
            task, env, expected = item
            claimed += 1
            try:
                self.run_eval_task(task, env, expected)
            except Exception as exc:
                print(f"[WORKER][FAIL][EVAL] {task.tag}: {exc}", flush=True)
            finally:
                eval_queue.task_done()
        print(f"[WORKER][DONE][EVAL] slot={worker_idx} completed={claimed}", flush=True)

    def print_plan(self) -> None:
        payload = {
            "tasks": len(self.tasks),
            "gpus": self.gpu_ids,
            "mode": self.run_mode,
            "resume_infer": self.resume_infer,
            "judge": self.judge,
            "eval_workers": self.eval_workers,
            "datasets": self.datasets,
            "combos": [f"{train}:{eval_}" for train, eval_ in self.combos],
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2), flush=True)

    def run(self) -> int:
        self.print_plan()
        if self.args.plan_only:
            return 0
        task_queue: queue.Queue[Task] = queue.Queue()
        eval_queue: queue.Queue[tuple[Task, dict[str, str], int] | None] = queue.Queue()
        for task in self.tasks:
            task_queue.put(task)
        infer_worker_count = min(len(self.gpu_ids), len(self.tasks))
        eval_worker_count = self.eval_workers if self.run_mode in {"all", "eval"} else 0
        total_workers = infer_worker_count + eval_worker_count
        with ThreadPoolExecutor(max_workers=total_workers) as pool:
            futures = []
            for worker_idx in range(eval_worker_count):
                futures.append(pool.submit(self.run_eval_worker, worker_idx, eval_queue))
            for worker_idx in range(infer_worker_count):
                futures.append(pool.submit(self.run_infer_worker, worker_idx, self.gpu_ids[worker_idx], task_queue, eval_queue))
            task_queue.join()
            for _ in range(eval_worker_count):
                eval_queue.put(None)
            eval_queue.join()
            for future in futures:
                future.result()
        print("[DONE] llava15 cross replay sweep finished", flush=True)
        return 0

    @staticmethod
    def _timestamp() -> str:
        return datetime.now().strftime("%Y%m%d%H%M%S")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Dynamic LLaVA checkpoint cross-replay launcher.")
    parser.add_argument("--repo-root", required=True)
    parser.add_argument("--python-bin", required=True)
    parser.add_argument("--ckpt-root", required=True)
    parser.add_argument("--work-root", required=True)
    parser.add_argument("--model-name", required=True)
    parser.add_argument("--mode", choices=["all", "infer", "eval"], default="all")
    parser.add_argument("--judge", default="gpt-4o-mini")
    parser.add_argument("--nproc", type=int, default=8)
    parser.add_argument("--eval-workers", type=int, default=4)
    parser.add_argument("--batch-size", type=int, default=16)
    parser.add_argument("--gpu-ids", default="0,1,2,3,4,5,6,7")
    parser.add_argument("--datasets", default="")
    parser.add_argument("--resume-infer", action="store_true")
    parser.add_argument("--plan-only", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    runner = LlavaCrossReplayRunner(args)
    return runner.run()


if __name__ == "__main__":
    raise SystemExit(main())
